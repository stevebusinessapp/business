from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg, F
from django.utils import timezone
from django.core.serializers import serialize
from django.template.loader import render_to_string
import json
import re
from decimal import Decimal
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import csv
from django import forms

from .models import (
    InventoryItem, InventoryLayout, InventoryStatus, InventoryCustomField,
    InventoryTransaction, InventoryLog, InventoryExport, ImportedInventoryFile,
    InventoryTemplate,
    # Legacy models
    InventoryProduct, InventoryCategory
)
from .forms import (
    InventoryItemForm, InventoryLayoutForm, InventoryCustomFieldForm,
    StatusChangeForm, InventoryTransactionForm, InventorySearchForm,
    InventoryImportForm, InventoryExportForm, InventoryTemplateForm,
    LayoutColumnForm, StockAdjustmentForm,
    # Legacy forms
    InventoryProductForm, InventoryCategoryForm
)


@login_required
def inventory_dashboard(request):
    """Main inventory dashboard with dynamic statistics"""
    # Get user's default layout or create one
    layout, created = InventoryLayout.objects.get_or_create(
        user=request.user,
        is_default=True,
        defaults={
            'name': 'Default Layout',
            'columns': InventoryLayout().get_default_columns()
        }
    )
    
    # Get comprehensive statistics
    total_items = InventoryItem.objects.filter(user=request.user).count()
    active_items = InventoryItem.objects.filter(user=request.user, is_active=True).count()
    
    # Calculate total value
    total_value = 0
    if layout.supports_calculations():
        items_with_totals = InventoryItem.objects.filter(user=request.user)
        total_value = sum(item.total_value for item in items_with_totals)
    
    # Status statistics
    status_stats = {}
    for status in InventoryStatus.objects.filter(is_active=True):
        count = InventoryItem.objects.filter(user=request.user, status=status).count()
        status_stats[status.name] = {
            'count': count,
            'color': status.color,
            'display_name': status.display_name
        }
    
    # Low stock alerts (items with quantity <= 5)
    low_stock_items = InventoryItem.objects.filter(
        user=request.user,
        is_active=True
    ).filter(
        Q(data__quantity__lte=5) | Q(data__Quantity__lte=5)
    )[:10]
    
    # Recent activity
    recent_logs = InventoryLog.objects.filter(user=request.user)[:10]
    
    # Layout statistics
    user_layouts = InventoryLayout.objects.filter(user=request.user)
    
    context = {
        'layout': layout,
        'total_items': total_items,
        'active_items': active_items,
        'total_value': total_value,
        'status_stats': status_stats,
        'low_stock_items': low_stock_items,
        'recent_logs': recent_logs,
        'user_layouts': user_layouts,
    }
    
    return render(request, 'inventory/dashboard.html', context)


@login_required
def inventory_list(request):
    """List inventory items with filtering and search"""
    # Clear any cached data to ensure fresh results
    from django.core.cache import cache
    # Clear specific cache keys instead of using delete_pattern
    cache_keys_to_clear = [
        'inventory_list_cache',
        'inventory_dashboard_cache',
        'inventory_export_cache',
    ]
    for key in cache_keys_to_clear:
        cache.delete(key)
    
    # Get layout
    layout_id = request.GET.get('layout')
    if layout_id:
        layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
    else:
        layout = InventoryLayout.objects.filter(user=request.user, is_default=True).first()
        if not layout:
            layout, _ = InventoryLayout.objects.get_or_create(
                user=request.user,
                is_default=True,
                defaults={
                    'name': 'Default Layout',
                    'columns': InventoryLayout().get_default_columns()
                }
            )
    
    # Handle category filtering
    category_id = request.GET.get('category')
    category = None
    if category_id:
        try:
            category = InventoryCategory.objects.get(pk=category_id, user=request.user)
        except InventoryCategory.DoesNotExist:
            pass
    
    # Get items for this user
    items = InventoryItem.objects.filter(user=request.user)
    
    # Apply category filtering if specified
    if category:
        # Try multiple filtering methods to ensure we find items
        category_items = items.filter(data__category=category.name)
        if not category_items.exists():
            # Try case-insensitive search
            category_items = items.filter(data__category__icontains=category.name)
        if not category_items.exists():
            # Try different field name
            category_items = items.filter(data__Category=category.name)
        
        items = category_items
    
    # Apply layout filtering AFTER category filtering (only if no category is specified)
    if layout and not category:
        items = items.filter(layout=layout)
    elif layout and category:
        # If both category and layout are specified, try to find items that match both
        layout_items = items.filter(layout=layout)
        if layout_items.exists():
            items = layout_items
        # If no items found with layout, show all items for the category regardless of layout
    
    # Force recalculation for all items to ensure latest data
    for item in items:
        item.calculate_totals()
    
    # Handle search and filtering
    search_form = InventorySearchForm(request.GET, user=request.user)
    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        status = search_form.cleaned_data.get('status')
        min_quantity = search_form.cleaned_data.get('min_quantity')
        max_quantity = search_form.cleaned_data.get('max_quantity')
        min_price = search_form.cleaned_data.get('min_price')
        max_price = search_form.cleaned_data.get('max_price')
        is_active = search_form.cleaned_data.get('is_active')
        
        if search:
            items = items.filter(
                Q(product_name__icontains=search) |
                Q(sku_code__icontains=search) |
                Q(data__icontains=search)
            )
        
        if status:
            items = items.filter(status=status)
        
        if min_quantity is not None:
            items = items.filter(
                Q(data__quantity__gte=min_quantity) | Q(data__Quantity__gte=min_quantity)
            )
        
        if max_quantity is not None:
            items = items.filter(
                Q(data__quantity__lte=max_quantity) | Q(data__Quantity__lte=max_quantity)
            )
        
        if min_price is not None:
            items = items.filter(
                Q(data__unit_price__gte=min_price) | Q(data__Unit_Price__gte=min_price)
            )
        
        if max_price is not None:
            items = items.filter(
                Q(data__unit_price__lte=max_price) | Q(data__Unit_Price__lte=max_price)
            )
        
        if is_active:
            items = items.filter(is_active=(is_active == 'true'))
    
    # Pagination
    paginator = Paginator(items, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate grand total if layout supports calculations
    grand_total = 0
    if layout.supports_calculations():
        grand_total = sum(item.total_value for item in page_obj)
    
    # Serialize layout data for JavaScript
    try:
        layout_data = {
            'id': layout.id,
            'name': layout.name,
            'columns': layout.columns if isinstance(layout.columns, (list, dict)) else [],
            'supports_calculations': layout.supports_calculations(),
            'calculation_fields': layout.get_calculation_fields(),
            'primary_color': str(layout.primary_color) if layout.primary_color else '#007bff',
            'secondary_color': str(layout.secondary_color) if layout.secondary_color else '#6c757d',
            'is_default': bool(layout.is_default),
        }
        layout_json = json.dumps(layout_data, ensure_ascii=False)
    except Exception as e:
        print(f"Error serializing layout data: {e}")
        # Fallback to basic data
        layout_data = {
            'id': layout.id,
            'name': layout.name,
            'columns': [],
            'supports_calculations': False,
            'calculation_fields': [],
            'primary_color': '#007bff',
            'secondary_color': '#6c757d',
            'is_default': False,
        }
        layout_json = json.dumps(layout_data, ensure_ascii=False)
    
    context = {
        'layout': layout,
        'layout_json': layout_json,
        'items': page_obj,
        'search_form': search_form,
        'grand_total': grand_total,
        'statuses': InventoryStatus.objects.filter(is_active=True),
        'user_layouts': InventoryLayout.objects.filter(user=request.user),
        'total_items': items.count(),
        'active_items': items.filter(is_active=True).count(),
        'low_stock_items': items.filter(
            Q(data__quantity__lte=5) | Q(data__Quantity__lte=5)
        ).count(),
        'supports_calculations': layout.supports_calculations(),
        'calculation_fields': layout.get_calculation_fields(),
        'current_category': category if category_id else None,
    }
    
    return render(request, 'inventory/inventory_list.html', context)


@login_required
def inventory_create(request):
    """Create a new inventory item"""
    # Get layout
    layout_id = request.GET.get('layout')
    if layout_id:
        layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
    else:
        layout = InventoryLayout.objects.filter(user=request.user, is_default=True).first()
        if not layout:
            layout, _ = InventoryLayout.objects.get_or_create(
                user=request.user,
                is_default=True,
                defaults={
                    'name': 'Default Layout',
                    'columns': InventoryLayout().get_default_columns()
                }
            )
    
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, user=request.user, layout=layout)
        
        if form.is_valid():
            item = form.save(commit=False)
            item.user = request.user
            item.layout = layout
            item.save()
            
            # Trigger updates across all documents and templates
            item.update_all_documents()
            
            # Log the creation
            InventoryLog.objects.create(
                user=request.user,
                item=item,
                log_type='create',
                description=f'Created item: {item.product_name}'
            )
            
            messages.success(request, f'Inventory item "{item.product_name}" created successfully! Quantity: {item.get_value("quantity")}, Unit Price: ₦{item.get_value("unit_price")}, Total: ₦{item.total_value}. You can now edit quantity and price directly in the table with real-time calculations.')
            
            return redirect('inventory:list')
    else:
        form = InventoryItemForm(user=request.user, layout=layout)
    
    context = {
        'form': form,
        'layout': layout,
        'title': 'Add New Item',
    }
    
    return render(request, 'inventory/inventory_form.html', context)


@login_required
def inventory_update(request, pk):
    """Update an inventory item"""
    item = get_object_or_404(InventoryItem, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item, user=request.user, layout=item.layout)
        if form.is_valid():
            old_data = {
                'product_name': item.product_name,
                'data': item.data.copy(),
            }
            
            item = form.save()
            
            # Trigger updates across all documents and templates
            item.update_all_documents()
            
            # Log the update
            changes = []
            if old_data['product_name'] != item.product_name:
                changes.append(f'Product name: {old_data["product_name"]} → {item.product_name}')
            
            # Check for data changes
            for key, new_value in item.data.items():
                old_value = old_data['data'].get(key)
                if old_value != new_value:
                    changes.append(f'{key}: {old_value} → {new_value}')
            
            if changes:
                InventoryLog.objects.create(
                    user=request.user,
                    item=item,
                    log_type='update',
                    description=f'Updated: {", ".join(changes)}'
                )
            
            messages.success(request, 'Inventory item updated successfully!')
            return redirect('inventory:list')
    else:
        form = InventoryItemForm(instance=item, user=request.user, layout=item.layout)
    
    context = {
        'form': form,
        'item': item,
        'layout': item.layout,
        'title': 'Edit Item',
    }
    
    return render(request, 'inventory/inventory_form.html', context)


@login_required
def inventory_delete(request, pk):
    """Delete an inventory item"""
    item = get_object_or_404(InventoryItem, pk=pk, user=request.user)
    
    if request.method == 'POST':
        item_name = item.product_name
        item.delete()
        
        messages.success(request, f'Item "{item_name}" deleted successfully!')
        return redirect('inventory:list')
    
    context = {
        'item': item,
    }
    
    return render(request, 'inventory/inventory_confirm_delete.html', context)


@login_required
def inventory_detail(request, pk):
    """View inventory item details"""
    from datetime import date, timedelta
    
    # Clear cache for this specific item to ensure fresh data
    from django.core.cache import cache
    cache.delete(f'inventory_item_{pk}')
    
    product = get_object_or_404(InventoryItem, pk=pk, user=request.user)
    
    # Force recalculation to ensure latest data
    product.calculate_totals()
    product.refresh_from_db()
    
    # Get recent activity logs
    recent_logs = InventoryLog.objects.filter(item=product).order_by('-created_at')[:10]
    
    # Date context for expiry date comparisons
    today = date.today()
    next_month = today + timedelta(days=30)
    
    # Process dynamic data for template - ensure all fields are available
    processed_data = {}
    
    # Add all data from the item's data field
    for field_name, value in product.data.items():
        processed_data[field_name] = product.get_value(field_name)
    
    # Ensure key fields are always present with defaults
    if 'category' not in processed_data:
        processed_data['category'] = None
    if 'supplier' not in processed_data:
        processed_data['supplier'] = None
    if 'location' not in processed_data:
        processed_data['location'] = None
    if 'minimum_threshold' not in processed_data:
        processed_data['minimum_threshold'] = None
    if 'description' not in processed_data:
        processed_data['description'] = None
    if 'notes' not in processed_data:
        processed_data['notes'] = None
    if 'expiry_date' not in processed_data:
        processed_data['expiry_date'] = None
    
    context = {
        'product': product,
        'layout': product.layout,
        'recent_logs': recent_logs,
        'today': today,
        'next_month': next_month,
        'processed_data': processed_data,
        'supports_calculations': product.layout.supports_calculations(),
        'calculated_data': product.calculated_data,
    }
    
    return render(request, 'inventory/inventory_detail.html', context)


# AJAX endpoints for real-time updates
@csrf_exempt
@require_POST
@login_required
def ajax_update_field(request):
    """Update a field value via AJAX with real-time calculations"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        field_name = data.get('field_name')
        value = data.get('value')
        field_type = data.get('field_type', 'text')
        
        item = get_object_or_404(InventoryItem, pk=item_id, user=request.user)
        
        # Sanitize and validate the value
        if field_type in ['number', 'decimal']:
            sanitized_value = item._extract_number(value)
            if sanitized_value is None:
                return JsonResponse({
                    'success': False,
                    'error': 'Please enter a valid number'
                })
            value = sanitized_value
        
        # Update the field
        item.set_value(field_name, value)
        
        # Force recalculation of totals if this is a calculation-related field
        calculated_data = {}
        if item.layout.supports_calculations():
            # Always recalculate when quantity or price fields change
            if field_name.lower() in ['quantity', 'quantity_in_stock', 'unit_price', 'price']:
                calculated_data = item.calculate_totals()
            else:
                calculated_data = item.calculate_totals()
        
        # Trigger updates across all documents and templates
        item.update_all_documents()
        
        # Create transaction record - handle serialization
        old_value = item.get_value(field_name)
        # Convert various types for JSON serialization
        if hasattr(old_value, 'as_tuple'):  # Check if it's a Decimal
            old_value = float(old_value)
        elif hasattr(old_value, 'isoformat'):  # Check if it's a date/datetime
            old_value = old_value.isoformat()
        if hasattr(value, 'as_tuple'):  # Check if it's a Decimal
            value = float(value)
        elif hasattr(value, 'isoformat'):  # Check if it's a date/datetime
            value = value.isoformat()
            
        InventoryTransaction.objects.create(
            user=request.user,
            item=item,
            transaction_type='field_update',
            field_changes={field_name: {'old': old_value, 'new': value}},
            notes=f'Field {field_name} updated via inline editing'
        )
        
        # Log the field update - handle Decimal serialization
        old_log_value = item.get_value(field_name)
        new_log_value = value
        
        # Convert Decimal objects to float for JSON serialization
        if hasattr(old_log_value, 'as_tuple'):  # Check if it's a Decimal
            old_log_value = float(old_log_value)
        if hasattr(new_log_value, 'as_tuple'):  # Check if it's a Decimal
            new_log_value = float(new_log_value)
            
        InventoryLog.objects.create(
            user=request.user,
            item=item,
            log_type='field_update',
            description=f'Updated {field_name}: {value}',
            details={'field_name': field_name, 'old_value': old_log_value, 'new_value': new_log_value}
        )
        
        return JsonResponse({
            'success': True,
            'new_value': str(value),
            'formatted_value': item.format_value(value, field_type),
            'total': str(item.total_value),
            'calculated_data': calculated_data,
            'message': f'{field_name} updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_update_status(request):
    """Update status via AJAX with enhanced tracking"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        status_id = data.get('new_status')  # Changed from 'status_id' to 'new_status'
        
        print(f"DEBUG: Updating status for item {item_id} to status {status_id}")
        
        item = get_object_or_404(InventoryItem, pk=item_id, user=request.user)
        
        # Get status by ID
        status = get_object_or_404(InventoryStatus, pk=status_id)
        
        old_status = item.status
        print(f"DEBUG: Old status - {old_status.name} ({old_status.display_name})")
        print(f"DEBUG: New status - {status.name} ({status.display_name})")
        
        item.status = status
        item.save()
        
        # Trigger updates across all documents and templates (skip automatic status update)
        item.update_all_documents(skip_status_update=True)
        
        # Force refresh of all related data
        from django.core.cache import cache
        cache_keys_to_clear = [
            f'inventory_item_{item.id}',
            f'inventory_user_{item.user.id}',
            f'inventory_layout_{item.layout.id}',
            f'inventory_status_{item.status.id}',
            f'inventory_status_{old_status.id}',
            f'inventory_category_{item.data.get("category", "")}',
            'inventory_list_cache',
            'inventory_dashboard_cache',
            'inventory_export_cache',
        ]
        
        for key in cache_keys_to_clear:
            cache.delete(key)
        
        print(f"DEBUG: Status saved successfully for item {item_id}")
        print(f"DEBUG: Cleared cache keys for fresh data across all views")
        
        # Create transaction record
        InventoryTransaction.objects.create(
            user=request.user,
            item=item,
            transaction_type='status_change',
            status_before=old_status,
            status_after=status,
            notes=f'Status changed from {old_status.display_name} to {status.display_name}'
        )
        
        # Log the status change
        InventoryLog.objects.create(
            user=request.user,
            item=item,
            log_type='status_change',
            description=f'Status changed: {old_status.display_name} → {status.display_name}',
            details={
                'old_status': old_status.name,
                'new_status': status.name,
                'old_display_name': old_status.display_name,
                'new_display_name': status.display_name
            }
        )
        
        return JsonResponse({
            'success': True,
            'status_name': status.name,
            'status_display_name': status.display_name,
            'status_color': status.color,
            'previous_status': old_status.name,
            'message': f'Status updated to {status.display_name}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_stock_adjustment(request):
    """Adjust stock quantity via AJAX"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        adjustment_type = data.get('adjustment_type')  # 'add', 'subtract', 'set'
        quantity = extract_numeric_value(data.get('quantity', 0))
        reason = data.get('reason', '')
        
        item = get_object_or_404(InventoryItem, pk=item_id, user=request.user)
        
        old_quantity = item.quantity
        
        if adjustment_type == 'add':
            new_quantity = old_quantity + quantity
        elif adjustment_type == 'subtract':
            new_quantity = max(0, old_quantity - quantity)
        elif adjustment_type == 'set':
            new_quantity = max(0, quantity)
        else:
            raise ValueError('Invalid adjustment type')
        
        # Update quantity
        item.set_value('quantity', new_quantity)
        
        # Update all related documents and calculations
        item.update_all_documents()
        
        # Create transaction log
        InventoryTransaction.objects.create(
            user=request.user,
            item=item,
            transaction_type='adjustment',
            quantity_change=new_quantity - old_quantity,
            quantity_before=old_quantity,
            quantity_after=new_quantity,
            notes=reason
        )
        
        # Log the adjustment
        InventoryLog.objects.create(
            user=request.user,
            item=item,
            log_type='stock_adjustment',
            description=f'Stock adjustment: {adjustment_type} {quantity} ({reason})'
        )
        
        return JsonResponse({
            'success': True,
            'new_quantity': new_quantity,
            'total': str(item.total_value),
            'message': f'Stock adjusted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_calculate_totals(request):
    """Calculate totals for all items in a layout with enhanced functionality"""
    try:
        data = json.loads(request.body)
        layout_id = data.get('layout_id')
        item_ids = data.get('item_ids', [])
        
        if layout_id:
            layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
            items = InventoryItem.objects.filter(user=request.user, layout=layout)
        elif item_ids:
            items = InventoryItem.objects.filter(user=request.user, pk__in=item_ids)
        else:
            items = InventoryItem.objects.filter(user=request.user)
        
        totals = {}
        grand_total = 0
        item_count = 0
        
        for item in items:
            calculated = item.calculate_totals()
            totals[item.id] = calculated
            
            # Add to grand total if calculations are supported
            if item.layout.supports_calculations():
                total_value = calculated.get('total', 0)
                grand_total += total_value
                item_count += 1
        
        return JsonResponse({
            'success': True,
            'totals': totals,
            'grand_total': grand_total,
            'formatted_grand_total': f"₦{grand_total:,.2f}",
            'item_count': item_count,
            'supports_calculations': any(item.layout.supports_calculations() for item in items)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# Layout management views
@login_required
def layout_list(request):
    """List user's layouts"""
    layouts = InventoryLayout.objects.filter(user=request.user)
    
    context = {
        'layouts': layouts,
    }
    
    return render(request, 'inventory/layout_list.html', context)


@login_required
def layout_create(request):
    """Create a new layout"""
    if request.method == 'POST':
        form = InventoryLayoutForm(request.POST, user=request.user)
        if form.is_valid():
            layout = form.save(commit=False)
            layout.user = request.user
            
            # Set default columns if not provided
            if not layout.columns:
                layout.columns = layout.get_default_columns()
            
            layout.save()
            
            messages.success(request, 'Layout created successfully!')
            return redirect('inventory:layout_list')
    else:
        form = InventoryLayoutForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Create Layout',
    }
    
    return render(request, 'inventory/layout_form.html', context)


@login_required
def layout_update(request, pk):
    """Update a layout"""
    layout = get_object_or_404(InventoryLayout, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = InventoryLayoutForm(request.POST, instance=layout, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Layout updated successfully!')
            return redirect('inventory:layout_list')
    else:
        form = InventoryLayoutForm(instance=layout, user=request.user)
    
    context = {
        'form': form,
        'layout': layout,
        'title': 'Edit Layout',
    }
    
    return render(request, 'inventory/layout_form.html', context)


@login_required
def layout_delete(request, pk):
    """Delete a layout"""
    layout = get_object_or_404(InventoryLayout, pk=pk, user=request.user)
    
    if request.method == 'POST':
        layout_name = layout.name
        layout.delete()
        
        messages.success(request, f'Layout "{layout_name}" deleted successfully!')
        return redirect('inventory:layout_list')
    
    context = {
        'layout': layout,
    }
    
    return render(request, 'inventory/layout_confirm_delete.html', context)


@login_required
def layout_detail(request, pk):
    """View layout details"""
    layout = get_object_or_404(InventoryLayout, pk=pk, user=request.user)
    
    context = {
        'layout': layout,
    }
    
    return render(request, 'inventory/layout_detail.html', context)


# Import/Export views
@login_required
def inventory_import(request):
    """Import inventory from Excel/CSV with smart column detection"""
    if request.method == 'POST':
        form = InventoryImportForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            try:
                uploaded_file = request.FILES['file']
                layout = form.cleaned_data.get('layout')
                
                # Determine file type
                file_extension = uploaded_file.name.split('.')[-1].lower()
                if file_extension in ['xlsx', 'xls']:
                    file_type = 'excel'
                elif file_extension == 'csv':
                    file_type = 'csv'
                else:
                    messages.error(request, 'Unsupported file format. Please use Excel (.xlsx, .xls) or CSV files.')
                    return redirect('inventory:import')
                
                # Read the file
                if file_type == 'excel':
                    df = pd.read_excel(uploaded_file)
                else:
                    df = pd.read_csv(uploaded_file)
                
                # Auto-detect columns and create mapping
                column_mapping = {}
                detected_columns = []
                
                for col in df.columns:
                    col_lower = col.lower().strip()
                    if any(keyword in col_lower for keyword in ['product', 'name', 'item']):
                        column_mapping[col] = 'product_name'
                        detected_columns.append(col)
                    elif any(keyword in col_lower for keyword in ['sku', 'code', 'id']):
                        column_mapping[col] = 'sku_code'
                        detected_columns.append(col)
                    elif any(keyword in col_lower for keyword in ['quantity', 'qty', 'stock']):
                        column_mapping[col] = 'quantity'
                        detected_columns.append(col)
                    elif any(keyword in col_lower for keyword in ['price', 'cost', 'unit']):
                        column_mapping[col] = 'unit_price'
                        detected_columns.append(col)
                    elif any(keyword in col_lower for keyword in ['status']):
                        column_mapping[col] = 'status'
                        detected_columns.append(col)
                
                # Create import record
                import_record = ImportedInventoryFile.objects.create(
                    user=request.user,
                    layout=layout,
                    file_name=uploaded_file.name,
                    file_path=f'inventory/imports/{uploaded_file.name}',
                    file_size=uploaded_file.size,
                    file_type=file_type,
                    column_mapping=column_mapping,
                    total_rows=len(df),
                    status='processing'
                )
                
                # Process the data
                imported_count = 0
                failed_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        # Extract values using mapping
                        product_name = row.get(column_mapping.get('product_name', ''), '')
                        sku_code = row.get(column_mapping.get('sku_code', ''), '')
                        quantity = row.get(column_mapping.get('quantity', ''), 0)
                        unit_price = row.get(column_mapping.get('unit_price', ''), 0)
                        status_name = row.get(column_mapping.get('status', ''), 'in_stock')
                        
                        # Validate required fields
                        if not product_name or not sku_code:
                            failed_count += 1
                            errors.append(f"Row {index + 1}: Missing product name or SKU")
                            continue
                        
                        # Get or create status
                        status, _ = InventoryStatus.objects.get_or_create(
                            name=status_name,
                            defaults={'display_name': status_name.replace('_', ' ').title()}
                        )
                        
                        # Check if item already exists
                        existing_item = InventoryItem.objects.filter(
                            user=request.user,
                            sku_code=sku_code
                        ).first()
                        
                        if existing_item:
                            # Update existing item
                            existing_item.product_name = product_name
                            existing_item.set_value('quantity', quantity)
                            existing_item.set_value('unit_price', unit_price)
                            existing_item.status = status
                            existing_item.save()
                        else:
                            # Create new item
                            item = InventoryItem.objects.create(
                                user=request.user,
                                layout=layout,
                                product_name=product_name,
                                sku_code=sku_code,
                                status=status,
                                data={
                                    'quantity': quantity,
                                    'unit_price': unit_price
                                }
                            )
                        
                        imported_count += 1
                        
                    except Exception as e:
                        failed_count += 1
                        errors.append(f"Row {index + 1}: {str(e)}")
                
                # Update import record
                import_record.imported_rows = imported_count
                import_record.failed_rows = failed_count
                import_record.error_log = '\n'.join(errors)
                import_record.status = 'completed'
                import_record.completed_at = timezone.now()
                import_record.save()
                
                # Log the import
                InventoryLog.objects.create(
                    user=request.user,
                    layout=layout,
                    log_type='import',
                    description=f'Imported {imported_count} items from {uploaded_file.name}',
                    details={
                        'file_name': uploaded_file.name,
                        'imported_count': imported_count,
                        'failed_count': failed_count,
                        'total_rows': len(df)
                    }
                )
                
                if imported_count > 0:
                    messages.success(request, f'Successfully imported {imported_count} items!')
                if failed_count > 0:
                    messages.warning(request, f'{failed_count} rows failed to import. Check the import log for details.')
                
                return redirect('inventory:list')
                
            except Exception as e:
                messages.error(request, f'Import failed: {str(e)}')
                return redirect('inventory:import')
    else:
        form = InventoryImportForm(user=request.user)
    
    # Get recent imports
    recent_imports = ImportedInventoryFile.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    context = {
        'form': form,
        'recent_imports': recent_imports,
        'title': 'Import Inventory',
    }
    
    return render(request, 'inventory/inventory_import.html', context)


@login_required
def inventory_export(request):
    """Export inventory to Excel/PDF with branding and calculations"""
    if request.method == 'POST':
        form = InventoryExportForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                layout = form.cleaned_data.get('layout')
                export_format = form.cleaned_data.get('format')
                include_calculations = form.cleaned_data.get('include_calculations', True)
                include_branding = form.cleaned_data.get('include_branding', True)
                
                # Get items for export
                items = InventoryItem.objects.filter(user=request.user, layout=layout)
                
                # Apply filters
                if form.cleaned_data.get('category_filter'):
                    items = items.filter(category=form.cleaned_data['category_filter'])
                
                if form.cleaned_data.get('status_filter'):
                    items = items.filter(status=form.cleaned_data['status_filter'])
                
                if form.cleaned_data.get('search'):
                    search = form.cleaned_data['search']
                    items = items.filter(
                        Q(product_name__icontains=search) |
                        Q(sku_code__icontains=search) |
                        Q(description__icontains=search)
                    )
                
                if form.cleaned_data.get('min_quantity'):
                    items = items.filter(data__quantity_in_stock__gte=form.cleaned_data['min_quantity'])
                
                if form.cleaned_data.get('max_quantity'):
                    items = items.filter(data__quantity_in_stock__lte=form.cleaned_data['max_quantity'])
                
                if form.cleaned_data.get('min_price'):
                    items = items.filter(data__unit_price__gte=form.cleaned_data['min_price'])
                
                if form.cleaned_data.get('max_price'):
                    items = items.filter(data__unit_price__lte=form.cleaned_data['max_price'])
                
                if form.cleaned_data.get('date_from'):
                    items = items.filter(created_at__gte=form.cleaned_data['date_from'])
                
                if form.cleaned_data.get('date_to'):
                    items = items.filter(created_at__lte=form.cleaned_data['date_to'])
                
                # Handle low stock filter
                if not form.cleaned_data.get('include_low_stock', True):
                    items = items.exclude(
                        data__quantity_in_stock__lt=F('data__minimum_threshold')
                    )
                
                # Generate filename
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                filename = f"inventory_export_{layout.name.replace(' ', '_')}_{timestamp}"
                
                if export_format == 'excel':
                    return export_to_excel(items, layout, filename, include_calculations, include_branding)
                elif export_format == 'csv':
                    return export_to_csv(items, layout, filename, include_calculations)
                elif export_format == 'pdf':
                    return export_to_pdf(items, layout, filename, include_calculations, include_branding)
                else:
                    messages.error(request, 'Unsupported export format')
                    return redirect('inventory:export')
                    
            except Exception as e:
                messages.error(request, f'Export failed: {str(e)}')
                return redirect('inventory:export')
    else:
        form = InventoryExportForm(user=request.user)
    
    # Get recent exports
    recent_exports = InventoryExport.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Get default layout for initial context
    default_layout = InventoryLayout.objects.filter(user=request.user, is_default=True).first()
    
    # Calculate initial summary statistics
    if default_layout:
        total_items = InventoryItem.objects.filter(user=request.user, layout=default_layout).count()
        total_value = sum(item.total_value for item in InventoryItem.objects.filter(user=request.user, layout=default_layout))
        
        # Count unique categories from data field
        categories = 0
        category_names = set()
        for item in InventoryItem.objects.filter(user=request.user, layout=default_layout):
            category_data = item.data.get('category', {})
            if isinstance(category_data, dict) and category_data.get('name'):
                category_names.add(category_data['name'])
        categories = len(category_names)
        
        # Count low stock items
        low_stock_count = 0
        for item in InventoryItem.objects.filter(user=request.user, layout=default_layout):
            quantity = item.data.get('quantity_in_stock', 0)
            min_threshold = item.data.get('minimum_threshold', 0)
            if quantity < min_threshold:
                low_stock_count += 1
    else:
        total_items = 0
        total_value = 0
        categories = 0
        low_stock_count = 0
    
    context = {
        'form': form,
        'recent_exports': recent_exports,
        'title': 'Export Inventory',
        'total_products': total_items,
        'total_value': total_value,
        'categories': categories,
        'low_stock_count': low_stock_count,
        'default_layout': default_layout,
    }
    
    return render(request, 'inventory/inventory_export.html', context)


def export_to_excel(items, layout, filename, include_calculations=True, include_branding=True):
    """Export inventory to Excel with formatting and company branding"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    # Force recalculation for all items to ensure latest data
    for item in items:
        item.calculate_totals()
        item.refresh_from_db()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Get company profile for branding
    company_profile = None
    if include_branding:
        try:
            company_profile = layout.user.company_profile
        except:
            pass
    
    # Add branding header if requested
    if include_branding and company_profile:
        # Add company logo if available
        logo_path = None
        if company_profile.logo and os.path.exists(company_profile.logo.path):
            try:
                logo_path = company_profile.logo.path
                # Add logo to Excel (positioned at A1)
                img = XLImage(logo_path)
                img.width = 100
                img.height = 60
                ws.add_image(img, 'A1')
                # Adjust row height to accommodate logo
                ws.row_dimensions[1].height = 50
            except:
                pass
        
        # Company name and details
        company_name_cell = ws['B1'] if logo_path else ws['A1']
        company_name_cell.value = company_profile.company_name
        company_name_cell.font = Font(size=18, bold=True, color="2E86AB")
        company_name_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Company details
        details_row = 2
        company_details = []
        if company_profile.address:
            company_details.append(company_profile.address)
        if company_profile.phone:
            company_details.append(f"Phone: {company_profile.phone}")
        if company_profile.email:
            company_details.append(f"Email: {company_profile.email}")
        if company_profile.website:
            company_details.append(f"Website: {company_profile.website}")
        
        if company_details:
            details_cell = ws[f'B{details_row}'] if logo_path else ws[f'A{details_row}']
            details_cell.value = " | ".join(company_details)
            details_cell.font = Font(size=10, color="666666")
            details_cell.alignment = Alignment(horizontal='left', vertical='center')
            details_row += 1
        
        # Add export info
        export_info_row = details_row + 1
        export_info_cell = ws[f'B{export_info_row}'] if logo_path else ws[f'A{export_info_row}']
        export_info_cell.value = f"Inventory Export Report - Generated on {timezone.now().strftime('%B %d, %Y at %H:%M')}"
        export_info_cell.font = Font(size=12, bold=True, italic=True, color="2E86AB")
        export_info_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        row_offset = export_info_row + 2
    else:
        row_offset = 1
    
    # Add headers
    headers = []
    for column in layout.get_visible_columns():
        if column.get('name') == 'actions':
            continue
        if not include_calculations and column.get('name') == 'total':
            continue
        headers.append(column.get('display_name', column.get('name')))
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_offset, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add data with enhanced styling
    currency_symbol = company_profile.currency_symbol if company_profile else '₦'
    
    for row, item in enumerate(items, row_offset + 1):
        col = 1
        for column in layout.get_visible_columns():
            if column.get('name') == 'actions':
                continue
            if not include_calculations and column.get('name') == 'total':
                continue
            
            field_name = column.get('name')
            if field_name == 'serial_number':
                value = row - row_offset  # Serial number based on row position
            elif field_name == 'product_name':
                value = item.product_name
            elif field_name == 'sku_code':
                value = item.sku_code
            elif field_name == 'status':
                value = item.status.display_name
            elif field_name == 'total':
                # Use company currency symbol if available
                currency_symbol = company_profile.currency_symbol if company_profile else '₦'
                # Only apply fallback for problematic symbols, otherwise use the actual symbol
                if currency_symbol == '₦':
                    # Try to use the actual Naira symbol, fallback to 'N' only if needed
                    try:
                        # Test if the symbol can be properly encoded
                        test_str = f"{currency_symbol}1000.00"
                        test_str.encode('utf-8')
                        # If no error, use the actual symbol
                        currency_symbol = '₦'
                    except UnicodeEncodeError:
                        currency_symbol = 'N'  # Fallback only if encoding fails
                value = f"{currency_symbol}{item.total_value:,.2f}"
            elif field_name == 'unit_price':
                # Use company currency symbol if available
                currency_symbol = company_profile.currency_symbol if company_profile else '₦'
                # Only apply fallback for problematic symbols, otherwise use the actual symbol
                if currency_symbol == '₦':
                    # Try to use the actual Naira symbol, fallback to 'N' only if needed
                    try:
                        # Test if the symbol can be properly encoded
                        test_str = f"{currency_symbol}1000.00"
                        test_str.encode('utf-8')
                        # If no error, use the actual symbol
                        currency_symbol = '₦'
                    except UnicodeEncodeError:
                        currency_symbol = 'N'  # Fallback only if encoding fails
                unit_price = item.data.get('unit_price', 0)
                value = f"{currency_symbol}{unit_price:,.2f}"
            else:
                value = item.get_value(field_name)
            
            cell = ws.cell(row=row, column=col, value=value)
            
            # Apply alternating row colors
            if (row - row_offset) % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            # Format numbers and currency
            if field_name in ['quantity']:
                cell.number_format = '#,##0.00'
            elif field_name in ['unit_price', 'total']:
                cell.number_format = f'"{currency_symbol}"#,##0.00'
            elif field_name == 'serial_number':
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            
            # Add borders
            cell.border = Border(
                left=Side(style='thin', color="CCCCCC"),
                right=Side(style='thin', color="CCCCCC"),
                top=Side(style='thin', color="CCCCCC"),
                bottom=Side(style='thin', color="CCCCCC")
            )
            
            col += 1
    
    # Add grand total if calculations are included
    if include_calculations and layout.supports_calculations():
        total_row = row_offset + len(items) + 1
        
        # Style for total row
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        
        # Total label
        total_label_cell = ws.cell(row=total_row, column=1, value="Grand Total")
        total_label_cell.font = total_font
        total_label_cell.fill = total_fill
        
        # Total value
        total_value = sum(item.total_value for item in items)
        total_value_cell = ws.cell(row=total_row, column=len(headers), value=total_value)
        total_value_cell.font = total_font
        total_value_cell.fill = total_fill
        total_value_cell.number_format = f'"{currency_symbol}"#,##0.00'
        
        # Add borders to total row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Auto-adjust column widths with better logic
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    # Calculate length based on content type
                    if isinstance(cell.value, (int, float)):
                        length = len(f"{cell.value:,.2f}")
                    else:
                        length = len(str(cell.value))
                    max_length = max(max_length, length)
            except:
                pass
        
        # Set reasonable column widths
        if max_length < 10:
            adjusted_width = 12
        elif max_length < 20:
            adjusted_width = 18
        elif max_length < 30:
            adjusted_width = 25
        else:
            adjusted_width = min(max_length + 5, 50)
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary information at the bottom
    summary_row = total_row + 3 if include_calculations and layout.supports_calculations() else row_offset + len(items) + 3
    
    # Summary section
    ws.cell(row=summary_row, column=1, value="Summary Information").font = Font(bold=True, size=12, color="2E86AB")
    summary_row += 1
    
    ws.cell(row=summary_row, column=1, value="Total Items:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(items))
    summary_row += 1
    
    ws.cell(row=summary_row, column=1, value="Report Generated:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=timezone.now().strftime('%B %d, %Y at %H:%M'))
    summary_row += 1
    
    ws.cell(row=summary_row, column=1, value="Layout:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=layout.name)
    
    if include_calculations and layout.supports_calculations():
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Total Inventory Value:").font = Font(bold=True)
        total_value = sum(item.total_value for item in items)
        ws.cell(row=summary_row, column=2, value=total_value)
        ws.cell(row=summary_row, column=2).number_format = f'"{currency_symbol}"#,##0.00'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb.save(response)
    return response


def export_to_csv(items, layout, filename, include_calculations=True):
    """Export inventory to CSV"""
    import csv
    
    # Force recalculation for all items to ensure latest data
    for item in items:
        item.calculate_totals()
        item.refresh_from_db()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = []
    for column in layout.get_visible_columns():
        if column.get('name') == 'actions':
            continue
        if not include_calculations and column.get('name') == 'total':
            continue
        headers.append(column.get('display_name', column.get('name')))
    
    writer.writerow(headers)
    
    # Write data
    for index, item in enumerate(items, 1):
        row = []
        for column in layout.get_visible_columns():
            if column.get('name') == 'actions':
                continue
            if not include_calculations and column.get('name') == 'total':
                continue
            
            field_name = column.get('name')
            if field_name == 'serial_number':
                value = index  # Serial number based on position
            elif field_name == 'product_name':
                value = item.product_name
            elif field_name == 'sku_code':
                value = item.sku_code
            elif field_name == 'status':
                value = item.status.display_name
            elif field_name == 'total':
                value = item.total_value
            else:
                value = item.get_value(field_name)
            
            row.append(value)
        
        writer.writerow(row)
    
    return response


def export_to_pdf(items, layout, filename, include_calculations=True, include_branding=True):
    """Export inventory to PDF with branding and company details"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    
    def get_currency_symbol_for_pdf(currency_symbol):
        """Get the appropriate currency symbol for PDF display"""
        if not currency_symbol or currency_symbol.strip() == '':
            return '₦'  # Default fallback
        
        # Clean the symbol
        currency_symbol = currency_symbol.strip()
        
        # Use a more conservative approach - use text representations for better compatibility
        currency_mapping = {
            '₦': 'NGN ',  # Naira - use code instead of symbol
            '$': '$',     # Dollar - works well
            '€': 'EUR ',  # Euro - use code instead of symbol
            '£': 'GBP ',  # Pound - use code instead of symbol
            '¥': 'JPY ',  # Yen - use code instead of symbol
            '₹': 'INR ',  # Rupee - use code instead of symbol
            '₽': 'RUB ',  # Ruble - use code instead of symbol
            '₩': 'KRW ',  # Won - use code instead of symbol
            '₪': 'ILS ',  # Shekel - use code instead of symbol
            '₨': 'PKR ',  # Rupee (Pakistani) - use code instead of symbol
            '₴': 'UAH ',  # Hryvnia - use code instead of symbol
            '₸': 'KZT ',  # Tenge - use code instead of symbol
            '₺': 'TRY ',  # Lira - use code instead of symbol
            '₼': 'AZN ',  # Manat - use code instead of symbol
            '₾': 'GEL ',  # Lari - use code instead of symbol
            '₿': 'BTC ',  # Bitcoin - use code instead of symbol
        }
        
        # If the symbol is in our mapping, use the text representation
        if currency_symbol in currency_mapping:
            return currency_mapping[currency_symbol]
        
        # For currency codes (like 'USD', 'EUR', etc.), add a space for readability
        if len(currency_symbol) == 3 and currency_symbol.isupper():
            return f"{currency_symbol} "
        
        # For unknown symbols, try to use them directly
        # If they fail, we'll fall back to a text representation
        try:
            # Test if the symbol can be properly encoded
            test_str = f"{currency_symbol}1000.00"
            test_str.encode('utf-8')
            # For safety, use a space after the symbol
            return f"{currency_symbol} "
        except UnicodeEncodeError:
            # If encoding fails, use a text fallback
            return f"{currency_symbol} "  # Add space for readability
    
    def wrap_text_for_pdf(text, max_chars_per_line=30):
        """Wrap text for better PDF display with more aggressive wrapping"""
        if not text:
            return ""
        
        text = str(text)
        if len(text) <= max_chars_per_line:
            return text
        
        # Try to break at spaces first
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            # If adding this word would exceed the limit
            if len(current_line + " " + word) > max_chars_per_line:
                if current_line:
                    lines.append(current_line.strip())
                    current_line = word
                else:
                    # Word is too long, break it more aggressively
                    if len(word) > max_chars_per_line:
                        # Break long words more frequently
                        for i in range(0, len(word), max_chars_per_line - 5):
                            lines.append(word[i:i + max_chars_per_line - 5])
                    else:
                        current_line = word
            else:
                if current_line:
                    current_line += " " + word
                else:
                    current_line = word
        
        if current_line:
            lines.append(current_line.strip())
        
        return "<br/>".join(lines)
    
    # Force recalculation for all items to ensure latest data
    for item in items:
        item.calculate_totals()
        item.refresh_from_db()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    # Use A4 landscape with optimized margins for better table layout
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Get company profile for branding
    company_profile = None
    if include_branding:
        try:
            company_profile = layout.user.company_profile
        except:
            pass
    
    # Add company header with logo
    if include_branding and company_profile:
        # Create header table with logo and company details
        header_data = []
        
        # Add logo if available
        logo_path = None
        if company_profile.logo and os.path.exists(company_profile.logo.path):
            try:
                logo_path = company_profile.logo.path
            except:
                pass
        
        if logo_path:
            # Header with logo and company name
            header_data.append([
                Image(logo_path, width=1.5*inch, height=1*inch),
                Paragraph(f"<b>{company_profile.company_name}</b>", 
                         ParagraphStyle('CompanyName', fontSize=16, alignment=TA_CENTER, spaceAfter=10))
            ])
        else:
            # Header with just company name
            header_data.append([
                "",
                Paragraph(f"<b>{company_profile.company_name}</b>", 
                         ParagraphStyle('CompanyName', fontSize=18, alignment=TA_CENTER, spaceAfter=10))
            ])
        
        # Add company details
        company_details = []
        if company_profile.address:
            company_details.append(company_profile.address)
        if company_profile.phone:
            company_details.append(f"Phone: {company_profile.phone}")
        if company_profile.email:
            company_details.append(f"Email: {company_profile.email}")
        if company_profile.website:
            company_details.append(f"Website: {company_profile.website}")
        
        if company_details:
            details_text = "<br/>".join(company_details)
            header_data.append([
                "",
                Paragraph(details_text, 
                         ParagraphStyle('CompanyDetails', fontSize=10, alignment=TA_CENTER, spaceAfter=20))
            ])
        
        # Create header table
        header_table = Table(header_data, colWidths=[2*inch, 4*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        story.append(header_table)
        story.append(Spacer(1, 20))
    
    # Add export info
    export_info = ParagraphStyle(
        'ExportInfo',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    story.append(Paragraph(f"<b>Inventory Export Report</b><br/>Generated on {timezone.now().strftime('%B %d, %Y at %H:%M')}", export_info))
    story.append(Spacer(1, 20))
    
    # Prepare table data with better column structure
    headers = []
    column_widths = []
    
    for column in layout.get_visible_columns():
        if column.get('name') == 'actions':
            continue
        if not include_calculations and column.get('name') == 'total':
            continue
        
        display_name = column.get('display_name', column.get('name'))
        headers.append(display_name)
        
        # Set appropriate column widths based on content type with better word wrapping support
        field_name = column.get('name')
        if field_name == 'serial_number':
            column_widths.append(0.35*inch)  # Compact but readable for serial numbers
        elif field_name == 'product_name':
            column_widths.append(1.8*inch)   # Balanced width for product names
        elif field_name == 'sku_code':
            column_widths.append(1.2*inch)   # Compact for SKU codes
        elif field_name == 'status':
            column_widths.append(0.8*inch)   # Compact for status
        elif field_name == 'quantity':
            column_widths.append(0.8*inch)   # Compact for quantities
        elif field_name == 'unit_price':
            column_widths.append(1.0*inch)   # Balanced for prices
        elif field_name == 'total':
            column_widths.append(1.4*inch)   # Wider width for totals to accommodate large amounts
        else:
            column_widths.append(0.9*inch)   # Compact default width for other fields
    
    table_data = [headers]
    
    # Define styles for word wrapping with better configuration
    cell_style = ParagraphStyle(
        'CellStyle',
        fontName='Helvetica',
        fontSize=8,  # Smaller font size for better fit
        leading=9,   # Reduced line spacing for better fit
        spaceBefore=0,
        spaceAfter=0,
        alignment=0,  # Left align
        wordWrap='LTR',  # Left-to-right word wrapping
        splitLongWords=True,  # Split long words if necessary
        spaceShrinkage=0.1,  # Allow more space reduction
    )
    
    for index, item in enumerate(items, 1):
        row = []
        for column in layout.get_visible_columns():
            if column.get('name') == 'actions':
                continue
            if not include_calculations and column.get('name') == 'total':
                continue
            
            field_name = column.get('name')
            if field_name == 'serial_number':
                value = str(index)  # Serial number based on position
            elif field_name == 'product_name':
                # Use improved text wrapping for product names
                wrapped_text = wrap_text_for_pdf(item.product_name, max_chars_per_line=25)
                value = Paragraph(wrapped_text, cell_style)
            elif field_name == 'sku_code':
                # Use improved text wrapping for SKU codes
                wrapped_text = wrap_text_for_pdf(item.sku_code, max_chars_per_line=20)
                value = Paragraph(wrapped_text, cell_style)
            elif field_name == 'status':
                # Use improved text wrapping for status
                status_text = item.status.display_name if item.status else 'Unknown'
                wrapped_text = wrap_text_for_pdf(status_text, max_chars_per_line=15)
                value = Paragraph(wrapped_text, cell_style)
            elif field_name == 'total':
                # Use company currency symbol if available
                raw_currency_symbol = company_profile.currency_symbol if company_profile else '₦'
                currency_symbol = get_currency_symbol_for_pdf(raw_currency_symbol)
                value = f"{currency_symbol}{item.total_value:,.2f}"
            elif field_name == 'unit_price':
                # Use company currency symbol if available
                raw_currency_symbol = company_profile.currency_symbol if company_profile else '₦'
                currency_symbol = get_currency_symbol_for_pdf(raw_currency_symbol)
                unit_price = item.data.get('unit_price', 0)
                value = f"{currency_symbol}{unit_price:,.2f}"
            else:
                # For other fields, use improved text wrapping
                field_value = item.get_value(field_name)
                wrapped_text = wrap_text_for_pdf(field_value, max_chars_per_line=20)
                value = Paragraph(wrapped_text, cell_style)
            
            row.append(value)
        
        table_data.append(row)
    
    # Add grand total if calculations are included
    if include_calculations and layout.supports_calculations():
        raw_currency_symbol = company_profile.currency_symbol if company_profile else '₦'
        currency_symbol = get_currency_symbol_for_pdf(raw_currency_symbol)
        total_value = sum(item.total_value for item in items)
        
        # Create grand total row with proper formatting
        total_row = []
        for i, column in enumerate(layout.get_visible_columns()):
            if column.get('name') == 'actions':
                continue
            if not include_calculations and column.get('name') == 'total':
                continue
            
            field_name = column.get('name')
            if field_name == 'total':
                # Format the grand total with proper currency symbol
                total_row.append(f"{currency_symbol}{total_value:,.2f}")
            elif field_name == 'serial_number':
                # Add "Grand Total" text in the first column
                total_row.append("Grand Total")
            else:
                # Empty cells for other columns
                total_row.append("")
        
        table_data.append(total_row)
    
    # Create table with proper column widths
    table = Table(table_data, colWidths=column_widths)
    
    # Enhanced table styling with word wrapping support
    table_style = [
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # Optimized padding for serial number header
        ('LEFTPADDING', (0, 0), (0, 0), 2),  # Balanced left padding for serial number header
        ('RIGHTPADDING', (0, 0), (0, 0), 2),  # Balanced right padding for serial number header
        
        # Extra padding for total column header
        ('LEFTPADDING', (-2, 0), (-2, 0), 3),  # Extra left padding for total column header
        ('RIGHTPADDING', (-2, 0), (-2, 0), 3),  # Extra right padding for total column header
        
        # Data rows styling with improved word wrapping support
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('VALIGN', (0, 1), (-1, -2), 'TOP'),  # TOP alignment for better word wrapping
        ('TOPPADDING', (0, 1), (-1, -2), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 2),
        ('LEFTPADDING', (0, 1), (-1, -2), 3),
        ('RIGHTPADDING', (0, 1), (-1, -2), 3),
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        
        # Alternating row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
        
        # Optimized padding for serial number column (first column)
        ('LEFTPADDING', (0, 1), (0, -2), 2),  # Balanced left padding for serial numbers
        ('RIGHTPADDING', (0, 1), (0, -2), 2),  # Balanced right padding for serial numbers
        
        # Extra padding for total column to accommodate large amounts
        ('LEFTPADDING', (-2, 1), (-2, -2), 3),  # Extra left padding for total column
        ('RIGHTPADDING', (-2, 1), (-2, -2), 3),  # Extra right padding for total column
    ]
    
    # Style the total row if it exists
    if include_calculations and layout.supports_calculations():
        table_style.extend([
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9ECEF')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('ALIGN', (0, -1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, -1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
            ('LEFTPADDING', (0, -1), (-1, -1), 2),
            ('RIGHTPADDING', (0, -1), (-1, -1), 2),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ])
    
    table.setStyle(TableStyle(table_style))
    
    story.append(table)
    
    # Add summary information (removed layout information)
    story.append(Spacer(1, 20))
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_LEFT
    )
    
    summary_text = f"""
    <b>Summary:</b><br/>
    • Total Items: {len(items)}<br/>
    • Report Generated: {timezone.now().strftime('%B %d, %Y at %H:%M')}<br/>
    """
    
    if include_calculations and layout.supports_calculations():
        raw_currency_symbol = company_profile.currency_symbol if company_profile else '₦'
        currency_symbol = get_currency_symbol_for_pdf(raw_currency_symbol)
        total_value = sum(item.total_value for item in items)
        summary_text += f"• Total Inventory Value: {currency_symbol}{total_value:,.2f}<br/>"
    
    story.append(Paragraph(summary_text, summary_style))
    
    doc.build(story)
    
    return response


@login_required
def stock_adjustment(request, pk):
    """Stock adjustment form"""
    product = get_object_or_404(InventoryItem, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            adjustment_type = form.cleaned_data['adjustment_type']
            quantity = form.cleaned_data['quantity']
            reason = form.cleaned_data['reason']
            notes = form.cleaned_data.get('notes', '')
            
            try:
                old_quantity = product.quantity
                
                # Calculate new quantity based on adjustment type
                if adjustment_type == 'add':
                    new_quantity = old_quantity + quantity
                elif adjustment_type == 'subtract':
                    new_quantity = max(0, old_quantity - quantity)
                elif adjustment_type == 'set':
                    new_quantity = max(0, quantity)
                else:
                    messages.error(request, 'Invalid adjustment type.')
                    return redirect('inventory:stock_adjustment', pk=pk)
                
                # Update the product quantity
                product.set_value('quantity', new_quantity)
                
                # Update all related documents and calculations
                product.update_all_documents()
                
                # Create transaction record
                InventoryTransaction.objects.create(
                    user=request.user,
                    item=product,
                    transaction_type='adjustment',
                    quantity_change=new_quantity - old_quantity,
                    quantity_before=old_quantity,
                    quantity_after=new_quantity,
                    notes=f"{reason}\n{notes}".strip()
                )
                
                # Log the adjustment
                InventoryLog.objects.create(
                    user=request.user,
                    item=product,
                    log_type='stock_adjustment',
                    description=f'Stock adjustment: {adjustment_type} {quantity} units. {reason}'
                )
                
                messages.success(request, f'Stock adjustment completed! Quantity updated from {old_quantity} to {new_quantity}.')
                return redirect('inventory:detail', pk=pk)
                
            except Exception as e:
                messages.error(request, f'Error processing adjustment: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    
    # Use the proper StockAdjustmentForm
    form = StockAdjustmentForm()
    
    context = {
        'form': form,
        'product': product,
        'title': 'Stock Adjustment',
    }
    
    return render(request, 'inventory/stock_adjustment.html', context)


# Utility functions
def extract_numeric_value(value):
    """Extract numeric value from mixed text/number input"""
    if not value:
        return 0
    
    # Convert to string and remove common non-numeric characters
    value_str = str(value)
    
    # Remove currency symbols, commas, and common text
    cleaned = re.sub(r'[^\d.-]', '', value_str)
    
    try:
        return Decimal(cleaned) if cleaned else 0
    except (ValueError, TypeError):
        return 0


# Legacy views for backward compatibility
@login_required
def product_list(request):
    """Legacy redirect to inventory list"""
    return redirect('inventory:list')


@login_required
def category_list(request):
    """List categories (legacy)"""
    categories = InventoryCategory.objects.filter(user=request.user)
    
    # Get company profile for currency information
    try:
        company_profile = request.user.company_profile
    except:
        company_profile = None
    
    # Calculate product count and total value for each category
    for category in categories:
        # Get items in this category - try multiple field variations
        category_items = InventoryItem.objects.filter(
            user=request.user,
            data__category=category.name
        )
        
        # If no items found, try case-insensitive search
        if category_items.count() == 0:
            category_items = InventoryItem.objects.filter(
                user=request.user,
                data__category__icontains=category.name
            )
        
        # If still no items, try different field names
        if category_items.count() == 0:
            category_items = InventoryItem.objects.filter(
                user=request.user,
                data__Category=category.name
            )
        
        # Calculate product count
        category.product_count = category_items.count()
        
        # Calculate total value
        category.total_value = 0
        for item in category_items:
            # Ensure item has total_value calculated
            if not hasattr(item, 'total_value') or item.total_value is None:
                item.calculate_totals()
            category.total_value += item.total_value
    
    context = {
        'categories': categories,
        'company_profile': company_profile,
    }
    
    return render(request, 'inventory/category_list.html', context)


@login_required
def category_create(request):
    """Create category (legacy)"""
    if request.method == 'POST':
        form = InventoryCategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            
            messages.success(request, 'Category created successfully!')
            return redirect('inventory:category_list')
    else:
        form = InventoryCategoryForm()
    
    context = {
        'form': form,
        'title': 'Add Category',
    }
    
    return render(request, 'inventory/category_form.html', context)


@login_required
def category_update(request, pk):
    """Update category (legacy)"""
    category = get_object_or_404(InventoryCategory, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = InventoryCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully!')
            return redirect('inventory:category_list')
    else:
        form = InventoryCategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
        'title': 'Edit Category',
    }
    
    return render(request, 'inventory/category_form.html', context)


# Template management views
@login_required
def template_list(request):
    """List all inventory templates"""
    templates = InventoryTemplate.objects.filter(
        Q(user=request.user) | Q(is_public=True)
    ).order_by('-created_at')
    
    return render(request, 'inventory/template_list.html', {
        'templates': templates,
        'title': 'Inventory Templates'
    })


@login_required
def template_create(request):
    """Create a new inventory template"""
    if request.method == 'POST':
        form = InventoryTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.user = request.user
            template.save()
            messages.success(request, 'Template created successfully.')
            return redirect('inventory:template_list')
    else:
        form = InventoryTemplateForm()
    
    return render(request, 'inventory/template_form.html', {
        'form': form,
        'title': 'Create Template'
    })


@login_required
def template_update(request, pk):
    """Update an inventory template"""
    template = get_object_or_404(InventoryTemplate, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = InventoryTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, 'Template updated successfully.')
            return redirect('inventory:template_list')
    else:
        form = InventoryTemplateForm(instance=template)
    
    return render(request, 'inventory/template_form.html', {
        'form': form,
        'template': template,
        'title': 'Update Template'
    })


@login_required
def template_delete(request, pk):
    """Delete an inventory template"""
    template = get_object_or_404(InventoryTemplate, pk=pk, user=request.user)
    
    if request.method == 'POST':
        template.delete()
        messages.success(request, 'Template deleted successfully.')
        return redirect('inventory:template_list')
    
    return render(request, 'inventory/template_confirm_delete.html', {
        'template': template,
        'title': 'Delete Template'
    })


@login_required
def template_detail(request, pk):
    """View template details"""
    template = get_object_or_404(
        InventoryTemplate, 
        pk=pk
    )
    
    # Check if user has access to this template
    if not (template.user == request.user or template.is_public):
        raise Http404("Template not found")
    
    return render(request, 'inventory/template_detail.html', {
        'template': template,
        'title': f'Template: {template.name}'
    }) 


@csrf_exempt
@require_POST
@login_required
def ajax_bulk_update_status(request):
    """Bulk update status for multiple items via AJAX"""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        new_status = data.get('new_status')
        
        if not item_ids:
            return JsonResponse({
                'success': False,
                'error': 'No items selected'
            })
        
        # Get status by name or ID
        if new_status.isdigit():
            status = get_object_or_404(InventoryStatus, pk=new_status)
        else:
            status = get_object_or_404(InventoryStatus, name=new_status)
        
        # Update all selected items
        items = InventoryItem.objects.filter(pk__in=item_ids, user=request.user)
        updated_count = 0
        
        for item in items:
            old_status = item.status
            item.status = status
            item.save()
            
            # Trigger updates across all documents and templates
            item.update_all_documents()
            
            # Create transaction record
            InventoryTransaction.objects.create(
                user=request.user,
                item=item,
                transaction_type='status_change',
                status_before=old_status,
                status_after=status,
                notes=f'Bulk status update to {status.display_name}'
            )
            
            # Log the status change
            InventoryLog.objects.create(
                user=request.user,
                item=item,
                log_type='status_change',
                description=f'Bulk status change: {old_status.display_name} → {status.display_name}',
                details={
                    'old_status': old_status.name,
                    'new_status': status.name,
                    'bulk_operation': True
                }
            )
            
            updated_count += 1
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'status_name': status.name,
            'status_display_name': status.display_name,
            'status_color': status.color,
            'message': f'Successfully updated {updated_count} items to {status.display_name}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_bulk_delete(request):
    """Bulk delete multiple items via AJAX"""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        
        if not item_ids:
            return JsonResponse({
                'success': False,
                'error': 'No items selected'
            })
        
        # Get items and delete them
        items = InventoryItem.objects.filter(pk__in=item_ids, user=request.user)
        deleted_count = 0
        
        for item in items:
            # Log before deletion
            InventoryLog.objects.create(
                user=request.user,
                item=item,
                log_type='delete',
                description=f'Bulk deleted: {item.product_name} ({item.sku_code})',
                details={
                    'product_name': item.product_name,
                    'sku_code': item.sku_code,
                    'bulk_operation': True
                }
            )
            
            item.delete()
            deleted_count += 1
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} items'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_GET
@login_required
def ajax_get_item_status(request, pk):
    """Get current item status via AJAX for real-time updates"""
    try:
        item = get_object_or_404(InventoryItem, pk=pk, user=request.user)
        
        # Force refresh from database to get latest status
        item.refresh_from_db()
        
        # Debug logging
        print(f"DEBUG: Item {pk} status - {item.status.name} ({item.status.display_name})")
        print(f"DEBUG: Item {pk} last updated - {item.updated_at}")
        
        return JsonResponse({
            'success': True,
            'status_name': item.status.name,
            'status_display_name': item.status.display_name,
            'status_color': item.status.color,
            'is_active': item.is_active,
            'quantity': item.quantity,
            'unit_price': float(item.unit_price),
            'total_value': float(item.total_value),
            'last_updated': item.updated_at.isoformat()
        })
        
    except Exception as e:
        print(f"ERROR in ajax_get_item_status: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_get_item_details(request):
    """Get detailed information about an item via AJAX"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        
        item = get_object_or_404(InventoryItem, pk=item_id, user=request.user)
        
        # Get all field values
        field_values = {}
        for column in item.layout.get_visible_columns():
            field_name = column.get('name')
            if field_name and field_name not in ['actions', 'total']:
                field_values[field_name] = item.get_value(field_name)
        
        # Get calculated data
        calculated_data = {}
        if item.layout.supports_calculations():
            calculated_data = item.calculate_totals()
        
        # Get recent transactions
        recent_transactions = item.transactions.order_by('-transaction_date')[:5]
        transaction_data = []
        for transaction in recent_transactions:
            transaction_data.append({
                'type': transaction.get_transaction_type_display(),
                'date': transaction.transaction_date.strftime('%Y-%m-%d %H:%M'),
                'notes': transaction.notes,
                'quantity_change': str(transaction.quantity_change) if transaction.quantity_change else None,
                'total_value': str(transaction.total_value) if transaction.total_value else None,
            })
        
        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'product_name': item.product_name,
                'sku_code': item.sku_code,
                'status': {
                    'name': item.status.name,
                    'display_name': item.status.display_name,
                    'color': item.status.color
                },
                'is_active': item.is_active,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': item.updated_at.strftime('%Y-%m-%d %H:%M'),
                'field_values': field_values,
                'calculated_data': calculated_data,
                'recent_transactions': transaction_data
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_quick_edit(request):
    """Quick edit multiple fields of an item via AJAX"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        field_updates = data.get('field_updates', {})
        
        item = get_object_or_404(InventoryItem, pk=item_id, user=request.user)
        
        # Track all changes
        changes = {}
        for field_name, new_value in field_updates.items():
            old_value = item.get_value(field_name)
            
            # Sanitize numeric values
            if field_name.lower() in ['quantity', 'unit_price', 'price', 'cost']:
                sanitized_value = item._extract_number(new_value)
                if sanitized_value is not None:
                    new_value = sanitized_value
            
            # Update the field
            item.set_value(field_name, new_value)
            
                    # Convert various types for JSON serialization
        if hasattr(old_value, 'as_tuple'):  # Check if it's a Decimal
            old_value = float(old_value)
        elif hasattr(old_value, 'isoformat'):  # Check if it's a date/datetime
            old_value = old_value.isoformat()
        if hasattr(new_value, 'as_tuple'):  # Check if it's a Decimal
            new_value = float(new_value)
        elif hasattr(new_value, 'isoformat'):  # Check if it's a date/datetime
            new_value = new_value.isoformat()
                
            changes[field_name] = {'old': old_value, 'new': new_value}
        
        # Recalculate totals if needed
        calculated_data = {}
        if item.layout.supports_calculations():
            calculated_data = item.calculate_totals()
        
        # Create transaction record
        InventoryTransaction.objects.create(
            user=request.user,
            item=item,
            transaction_type='field_update',
            field_changes=changes,
            notes=f'Quick edit: Updated {len(changes)} fields'
        )
        
        # Log the changes - changes dict already has Decimal objects converted to float
        InventoryLog.objects.create(
            user=request.user,
            item=item,
            log_type='field_update',
            description=f'Quick edit: Updated {len(changes)} fields',
            details={'changes': changes, 'quick_edit': True}
        )
        
        return JsonResponse({
            'success': True,
            'changes': changes,
            'calculated_data': calculated_data,
            'message': f'Successfully updated {len(changes)} fields'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_save_layout(request):
    """Save layout configuration via AJAX"""
    try:
        data = json.loads(request.body)
        layout_id = data.get('layout_id')
        layout_config = data.get('layout_config', {})
        
        if layout_id:
            layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
        else:
            layout = InventoryLayout.objects.create(
                user=request.user,
                name=data.get('name', 'New Layout'),
                description=data.get('description', '')
            )
        
        # Update layout configuration
        for key, value in layout_config.items():
            if hasattr(layout, key):
                setattr(layout, key, value)
        
        layout.save()
        
        return JsonResponse({
            'success': True,
            'layout_id': layout.id,
            'message': 'Layout saved successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_update_column_order(request):
    """Update column order via AJAX"""
    try:
        data = json.loads(request.body)
        layout_id = data.get('layout_id')
        column_order = data.get('column_order', [])
        
        layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
        layout.column_order = column_order
        layout.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Column order updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_toggle_column_visibility(request):
    """Toggle column visibility via AJAX"""
    try:
        data = json.loads(request.body)
        layout_id = data.get('layout_id')
        column_name = data.get('column_name')
        is_visible = data.get('is_visible', True)
        
        layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
        
        if not layout.column_visibility:
            layout.column_visibility = {}
        
        layout.column_visibility[column_name] = is_visible
        layout.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Column {column_name} visibility updated'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_validate_field(request):
    """Validate field value via AJAX"""
    try:
        data = json.loads(request.body)
        field_name = data.get('field_name')
        field_value = data.get('field_value')
        field_type = data.get('field_type', 'text')
        
        # Basic validation
        errors = []
        
        if field_type in ['number', 'decimal']:
            if field_value and not re.match(r'^-?\d*\.?\d+$', str(field_value)):
                errors.append('Please enter a valid number')
        
        elif field_type == 'email':
            if field_value and not re.match(r'^[^@]+@[^@]+\.[^@]+$', str(field_value)):
                errors.append('Please enter a valid email address')
        
        elif field_type == 'url':
            if field_value and not re.match(r'^https?://.+', str(field_value)):
                errors.append('Please enter a valid URL starting with http:// or https://')
        
        return JsonResponse({
            'success': True,
            'is_valid': len(errors) == 0,
            'errors': errors
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_get_calculation_preview(request):
    """Get calculation preview via AJAX"""
    try:
        data = json.loads(request.body)
        formula = data.get('formula', '')
        field_values = data.get('field_values', {})
        
        # Safe evaluation of formula
        result = None
        try:
            # Replace field names with values
            for field_name, value in field_values.items():
                formula = formula.replace(f'{{{field_name}}}', str(value))
            
            # Only allow basic math operations
            allowed_chars = set('0123456789+-*/(). ')
            if all(c in allowed_chars for c in formula):
                result = eval(formula)
            else:
                result = 'Invalid formula'
        except Exception:
            result = 'Calculation error'
        
        return JsonResponse({
            'success': True,
            'result': result
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_export_selected(request):
    """Export selected items via AJAX"""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        export_format = data.get('format', 'excel')
        include_calculations = data.get('include_calculations', True)
        include_branding = data.get('include_branding', True)
        
        # Get layout from first item
        if item_ids:
            first_item = InventoryItem.objects.get(pk=item_ids[0], user=request.user)
            layout = first_item.layout
        else:
            layout = InventoryLayout.objects.filter(user=request.user, is_default=True).first()
        
        # Get items
        items = InventoryItem.objects.filter(pk__in=item_ids, user=request.user)
        
        # Generate filename
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"inventory_export_{timestamp}"
        
        # Export based on format
        if export_format == 'excel':
            response = export_to_excel(items, layout, filename, include_calculations, include_branding)
        elif export_format == 'csv':
            response = export_to_csv(items, layout, filename, include_calculations)
        elif export_format == 'pdf':
            response = export_to_pdf(items, layout, filename, include_calculations, include_branding)
        else:
            return JsonResponse({
                'success': False,
                'error': 'Unsupported export format'
            })
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def inventory_print(request, pk):
    """Print view for inventory item"""
    item = get_object_or_404(InventoryItem, pk=pk, user=request.user)
    
    # Force recalculation to ensure latest data
    item.calculate_totals()
    item.refresh_from_db()
    
    context = {
        'item': item,
        'layout': item.layout,
        'supports_calculations': item.layout.supports_calculations(),
        'calculated_data': item.calculated_data,
    }
    
    return render(request, 'inventory/inventory_print.html', context)


@login_required
def custom_field_list(request):
    """List custom fields for the user"""
    custom_fields = InventoryCustomField.objects.filter(user=request.user).order_by('sort_order')
    
    context = {
        'custom_fields': custom_fields,
    }
    
    return render(request, 'inventory/custom_field_list.html', context)


@login_required
def custom_field_create(request):
    """Create a new custom field"""
    if request.method == 'POST':
        form = InventoryCustomFieldForm(request.POST, user=request.user)
        if form.is_valid():
            custom_field = form.save(commit=False)
            custom_field.user = request.user
            custom_field.save()
            
            messages.success(request, 'Custom field created successfully.')
            return redirect('inventory:custom_field_list')
    else:
        form = InventoryCustomFieldForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Create Custom Field'
    }
    
    return render(request, 'inventory/custom_field_form.html', context)


@login_required
def custom_field_update(request, pk):
    """Update a custom field"""
    custom_field = get_object_or_404(InventoryCustomField, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = InventoryCustomFieldForm(request.POST, instance=custom_field, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Custom field updated successfully.')
            return redirect('inventory:custom_field_list')
    else:
        form = InventoryCustomFieldForm(instance=custom_field, user=request.user)
    
    context = {
        'form': form,
        'custom_field': custom_field,
        'title': 'Update Custom Field'
    }
    
    return render(request, 'inventory/custom_field_form.html', context)


@login_required
def custom_field_delete(request, pk):
    """Delete a custom field"""
    custom_field = get_object_or_404(InventoryCustomField, pk=pk, user=request.user)
    
    if request.method == 'POST':
        custom_field.delete()
        messages.success(request, 'Custom field deleted successfully.')
        return redirect('inventory:custom_field_list')
    
    context = {
        'custom_field': custom_field,
    }
    
    return render(request, 'inventory/custom_field_confirm_delete.html', context) 


@csrf_exempt
@require_POST
@login_required
def ajax_set_default_layout(request):
    """Set a layout as default via AJAX"""
    try:
        data = json.loads(request.body)
        layout_id = data.get('layout_id')
        
        if not layout_id:
            return JsonResponse({
                'success': False,
                'error': 'Layout ID is required'
            })
        
        layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
        
        # Set this layout as default
        InventoryLayout.objects.filter(user=request.user, is_default=True).update(is_default=False)
        layout.is_default = True
        layout.save()
        
        # Log the change
        InventoryLog.objects.create(
            user=request.user,
            layout=layout,
            log_type='layout_change',
            description=f'Set "{layout.name}" as default layout',
            details={
                'action': 'set_default',
                'layout_id': layout.id,
                'layout_name': layout.name
            }
        )
        
        return JsonResponse({
            'success': True,
            'layout_name': layout.name,
            'message': f'"{layout.name}" is now your default layout'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_POST
@login_required
def ajax_layout_preview(request):
    """Generate layout preview for the form"""
    try:
        data = json.loads(request.body)
        form_data = data.get('form_data', {})
        
        # Create a mock layout for preview
        preview_layout = {
            'name': form_data.get('name', 'Preview Layout'),
            'primary_color': form_data.get('primary_color', '#007bff'),
            'secondary_color': form_data.get('secondary_color', '#6c757d'),
            'company_name': form_data.get('company_name', 'Your Company'),
            'supports_calculations': form_data.get('auto_calculate', True),
            'show_totals': form_data.get('show_totals', True),
            'show_grand_total': form_data.get('show_grand_total', True),
        }
        
        # Generate preview HTML
        preview_html = render_to_string('inventory/layout_preview.html', {
            'layout': preview_layout,
            'sample_items': [
                {
                    'product_name': 'Sample Product 1',
                    'sku_code': 'SKU001',
                    'quantity': 10,
                    'unit_price': 1500.00,
                    'total_value': 15000.00,
                    'status': {'display_name': 'In Stock', 'color': '#28a745'}
                },
                {
                    'product_name': 'Sample Product 2',
                    'sku_code': 'SKU002',
                    'quantity': 5,
                    'unit_price': 2500.00,
                    'total_value': 12500.00,
                    'status': {'display_name': 'Low Stock', 'color': '#ffc107'}
                }
            ]
        })
        
        return HttpResponse(preview_html)
        
    except Exception as e:
        return HttpResponse(f'<p class="text-danger">Error generating preview: {str(e)}</p>') 


@csrf_exempt
@require_POST
@login_required
def update_inventory_currency(request):
    """Update currency across all inventory documents and templates"""
    try:
        data = json.loads(request.body)
        currency_code = data.get('currency_code')
        currency_symbol = data.get('currency_symbol')
        
        if not currency_code or not currency_symbol:
            return JsonResponse({'success': False, 'error': 'Currency code and symbol are required'})
        
        # Update company profile currency
        company_profile = request.user.company_profile
        company_profile.currency_code = currency_code
        company_profile.currency_symbol = currency_symbol
        company_profile.save()
        
        # Update all inventory items to trigger recalculation
        inventory_items = InventoryItem.objects.filter(user=request.user)
        updated_count = 0
        
        for item in inventory_items:
            # Trigger recalculation to update currency displays
            item.calculate_totals()
            updated_count += 1
        
        # Update all inventory layouts
        layouts = InventoryLayout.objects.filter(user=request.user)
        for layout in layouts:
            layout.updated_at = timezone.now()
            layout.save()
        
        # Log the currency change
        InventoryLog.objects.create(
            user=request.user,
            log_type='field_update',
            description=f'Currency updated to {currency_code} ({currency_symbol})',
            details={
                'currency_code': currency_code,
                'currency_symbol': currency_symbol,
                'items_updated': updated_count,
                'layouts_updated': layouts.count(),
                'updated_at': timezone.now().isoformat()
            }
        )
        
        return JsonResponse({
            'success': True,
            'currency_code': currency_code,
            'currency_symbol': currency_symbol,
            'items_updated': updated_count,
            'layouts_updated': layouts.count(),
            'message': f'Currency updated successfully. {updated_count} items and {layouts.count()} layouts updated.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}) 


@csrf_exempt
@require_POST
@login_required
def ajax_export_preview(request):
    """Get export preview data via AJAX"""
    try:
        data = json.loads(request.body)
        
        # Get layout
        layout_id = data.get('layout_id')
        if layout_id:
            layout = get_object_or_404(InventoryLayout, pk=layout_id, user=request.user)
        else:
            layout = InventoryLayout.objects.filter(user=request.user, is_default=True).first()
            if not layout:
                return JsonResponse({
                    'success': False,
                    'error': 'No layout found'
                })
        
        # Get items with filters
        items = InventoryItem.objects.filter(user=request.user, layout=layout)
        
        # Apply filters
        if data.get('category_filter'):
            # Filter by category stored in data field
            items = items.filter(data__category__id=data['category_filter'])
        
        if data.get('status_filter'):
            items = items.filter(status__pk=data['status_filter'])
        
        if data.get('search'):
            search = data['search']
            items = items.filter(
                Q(product_name__icontains=search) |
                Q(sku_code__icontains=search) |
                Q(data__description__icontains=search)
            )
        
        if data.get('min_quantity'):
            items = items.filter(data__quantity_in_stock__gte=data['min_quantity'])
        
        if data.get('max_quantity'):
            items = items.filter(data__quantity_in_stock__lte=data['max_quantity'])
        
        if data.get('min_price'):
            items = items.filter(data__unit_price__gte=data['min_price'])
        
        if data.get('max_price'):
            items = items.filter(data__unit_price__lte=data['max_price'])
        
        if data.get('date_from'):
            items = items.filter(created_at__gte=data['date_from'])
        
        if data.get('date_to'):
            items = items.filter(created_at__lte=data['date_to'])
        
        # Handle low stock filter
        if not data.get('include_low_stock', True):
            # Exclude items below minimum threshold
            items = items.exclude(
                data__quantity_in_stock__lt=F('data__minimum_threshold')
            )
        
        # Calculate summary statistics
        total_items = items.count()
        total_value = sum(item.total_value for item in items)
        
        # Get unique categories from data field
        category_names = set()
        for item in items:
            category_data = item.data.get('category', {})
            if isinstance(category_data, dict) and category_data.get('name'):
                category_names.add(category_data['name'])
        categories = len(category_names)
        
        # Count low stock items
        low_stock_count = 0
        for item in items:
            quantity = item.data.get('quantity_in_stock', 0)
            min_threshold = item.data.get('minimum_threshold', 0)
            if quantity < min_threshold:
                low_stock_count += 1
        
        # Get preview items (first 10)
        preview_items = items[:10]
        
        # Prepare preview data
        preview_data = []
        for item in preview_items:
            category_data = item.data.get('category', {})
            category_name = category_data.get('name', 'Uncategorized') if isinstance(category_data, dict) else 'Uncategorized'
            
            preview_data.append({
                'id': item.pk,
                'product_name': item.product_name,
                'sku_code': item.sku_code,
                'category': category_name,
                'quantity': item.data.get('quantity_in_stock', 0),
                'unit_price': item.data.get('unit_price', 0),
                'total_value': item.total_value,
                'status': item.status.display_name if item.status else 'Active',
                'is_low_stock': item.data.get('quantity_in_stock', 0) < item.data.get('minimum_threshold', 0)
            })
        
        return JsonResponse({
            'success': True,
            'preview_data': preview_data,
            'summary': {
                'total_items': total_items,
                'total_value': float(total_value),
                'categories': categories,
                'low_stock_count': low_stock_count
            },
            'layout': {
                'name': layout.name,
                'supports_calculations': layout.supports_calculations()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })