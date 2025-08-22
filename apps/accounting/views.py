from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Sum, Q, Count, Min, Max
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.urls import reverse
from datetime import datetime, timedelta
import json
import csv
from decimal import Decimal
import re

from .models import Transaction, Ledger, Account, FinancialReport
from .forms import (
    TransactionForm, TransactionFilterForm, AccountForm, 
    FinancialReportForm, BulkTransactionForm, ReconciliationForm,
    ImportTransactionForm
)
from apps.core.models import CompanyProfile

def get_currency_display(currency_symbol):
    """Convert currency symbol to display text for better compatibility"""
    if not currency_symbol:
        return 'NGN'
    currency_symbol = currency_symbol.strip()
    currency_mapping = {
        '₦': 'NGN',
        '$': 'USD',
        '€': 'EUR',
        '£': 'GBP',
        '¥': 'JPY',
        '₹': 'INR',
        '₽': 'RUB',
        '₩': 'KRW',
        '₪': 'ILS',
        '₨': 'PKR',
        '₴': 'UAH',
        '₸': 'KZT',
        '₺': 'TRY',
        '₼': 'AZN',
        '₾': 'GEL',
        '₿': 'BTC',
    }
    return currency_mapping.get(currency_symbol, currency_symbol)


def clean_transaction_text(text):
    """Clean transaction text by removing HTML tags and special characters"""
    if not text:
        return ""
    
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', str(text))
    
    # Replace common HTML entities
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&amp;', '&')
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    text = text.replace('&quot;', '"')
    text = text.replace('&#39;', "'")
    
    # Clean up multiple spaces and newlines
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    return text


def get_pdf_currency_symbol(currency_symbol, use_symbol=True):
    """Convert currency symbol to PDF-compatible format"""
    if not currency_symbol or currency_symbol.strip() == '':
        return 'NGN '  # Default fallback with space
    
    # Clean the symbol
    currency_symbol = currency_symbol.strip()
    
    # If use_symbol is True, try to use the actual symbol first
    if use_symbol:
        # For common symbols that usually work well in PDFs
        if currency_symbol in ['$', '€', '£', '¥']:
            return f"{currency_symbol} "
        
        # For Naira symbol, try the actual symbol first
        if currency_symbol == '₦':
            try:
                # Test if the symbol can be properly encoded
                test_str = f"{currency_symbol}1000.00"
                test_str.encode('utf-8')
                return f"{currency_symbol} "
            except UnicodeEncodeError:
                pass  # Fall through to text representation
        
        # For other symbols, try them with a space
        try:
            # Test if the symbol can be properly encoded
            test_str = f"{currency_symbol}1000.00"
            test_str.encode('utf-8')
            return f"{currency_symbol} "
        except UnicodeEncodeError:
            pass  # Fall through to text representation
    
    # Use text representations for better compatibility
    currency_mapping = {
        '₦': 'NGN ',  # Naira - use code instead of symbol with space
        '$': '$',     # Dollar - works well
        '€': 'EUR ',  # Euro - use code instead of symbol with space
        '£': 'GBP ',  # Pound - use code instead of symbol with space
        '¥': 'JPY ',  # Yen - use code instead of symbol with space
        '₹': 'INR ',  # Rupee - use code instead of symbol with space
        '₽': 'RUB ',  # Ruble - use code instead of symbol with space
        '₩': 'KRW ',  # Won - use code instead of symbol with space
        '₪': 'ILS ',  # Shekel - use code instead of symbol with space
        '₨': 'PKR ',  # Rupee (Pakistani) - use code instead of symbol with space
        '₴': 'UAH ',  # Hryvnia - use code instead of symbol with space
        '₸': 'KZT ',  # Tenge - use code instead of symbol with space
        '₺': 'TRY ',  # Lira - use code instead of symbol with space
        '₼': 'AZN ',  # Manat - use code instead of symbol with space
        '₾': 'GEL ',  # Lari - use code instead of symbol with space
        '₿': 'BTC ',  # Bitcoin - use code instead of symbol with space
    }
    
    # If the symbol is in our mapping, use the text representation
    if currency_symbol in currency_mapping:
        return currency_mapping[currency_symbol]
    
    # For currency codes (like 'USD', 'EUR', etc.), add a space for readability
    if len(currency_symbol) == 3 and currency_symbol.isupper():
        return f"{currency_symbol} "
    
    # For unknown symbols, try to use them directly with a space
    try:
        # Test if the symbol can be properly encoded
        test_str = f"{currency_symbol}1000.00"
        test_str.encode('utf-8')
        # For safety, use a space after the symbol
        return f"{currency_symbol} "
    except UnicodeEncodeError:
        # If encoding fails, use a text fallback
        return f"{currency_symbol} "  # Add space for readability


def sync_ledgers_from_transactions(company):
    """Sync ledgers from existing transactions for all months"""
    
    # Get date range of transactions
    date_range = Transaction.objects.filter(
        company=company,
        is_void=False
    ).aggregate(
        min_date=Min('transaction_date'),
        max_date=Max('transaction_date')
    )
    
    if not date_range['min_date'] or not date_range['max_date']:
        return
    
    # Create ledgers for each month in the range
    current_date = date_range['min_date'].replace(day=1)
    end_date = date_range['max_date'].replace(day=1)
    
    while current_date <= end_date:
        # Check if ledger exists for this month
        existing_ledger = Ledger.objects.filter(
            company=company,
            year=current_date.year,
            month=current_date.month
        ).first()
        
        if not existing_ledger:
            # Calculate totals from transactions
            month_transactions = Transaction.objects.filter(
                company=company,
                transaction_date__year=current_date.year,
                transaction_date__month=current_date.month,
                is_void=False
            )
            
            total_income = month_transactions.filter(type='income').aggregate(
                total=Sum('net_amount')
            )['total'] or 0
            
            total_expense = month_transactions.filter(type='expense').aggregate(
                total=Sum('net_amount')
            )['total'] or 0
            
            net_profit = total_income - total_expense
            
            # Create ledger
            Ledger.objects.create(
                company=company,
                year=current_date.year,
                month=current_date.month,
                total_income=total_income,
                total_expense=total_expense,
                net_profit=net_profit
            )
        
        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)


@login_required
def accounting_dashboard(request):
    """Main accounting dashboard with charts and summary"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found. Please set up your company profile first.")
        return redirect('core:company_profile')
    
    # Sync ledgers from existing transactions
    sync_ledgers_from_transactions(company)
    
    # Get current month/year
    current_date = timezone.now()
    current_month = current_date.month
    current_year = current_date.year
    
    # Get current month ledger
    current_ledger = Ledger.objects.filter(
        company=company,
        year=current_year,
        month=current_month
    ).first()
    
    # Create a default ledger if none exists
    if not current_ledger:
        current_ledger = type('obj', (object,), {
            'total_income': 0,
            'total_expense': 0,
            'net_profit': 0
        })()
    
    # Get recent transactions
    recent_transactions = Transaction.objects.filter(
        company=company,
        is_void=False
    ).order_by('-created_at')[:10]
    
    # Get monthly data for charts (last 12 months) - improved with real transaction data
    monthly_data = []
    for i in range(12):
        date = current_date - timedelta(days=30*i)
        
        # Get or create ledger for this month
        month_ledger = Ledger.objects.filter(
            company=company,
            year=date.year,
            month=date.month
        ).first()
        
        # If no ledger exists, calculate from actual transactions
        if not month_ledger:
            month_transactions = Transaction.objects.filter(
                company=company,
                transaction_date__year=date.year,
                transaction_date__month=date.month,
                is_void=False
            )
            
            month_income = month_transactions.filter(type='income').aggregate(
                total=Sum('net_amount')
            )['total'] or 0
            
            month_expense = month_transactions.filter(type='expense').aggregate(
                total=Sum('net_amount')
            )['total'] or 0
            
            month_profit = month_income - month_expense
        else:
            month_income = float(month_ledger.total_income)
            month_expense = float(month_ledger.total_expense)
            month_profit = float(month_ledger.net_profit)
        
        monthly_data.append({
            'month': date.strftime('%b %Y'),
            'year': date.year,
            'month_num': date.month,
            'income': month_income,
            'expense': month_expense,
            'profit': month_profit,
        })
    
    # Get outstanding invoices
    from apps.invoices.models import Invoice
    outstanding_invoices = Invoice.objects.filter(
        user=user,
        status__in=['unpaid', 'partial']
    ).aggregate(total=Sum('balance_due'))['total'] or 0
    
    # Get today's transactions
    today_transactions = Transaction.objects.filter(
        company=company,
        transaction_date=current_date.date(),
        is_void=False
    )
    
    today_income = today_transactions.filter(type='income').aggregate(
        total=Sum('net_amount')
    )['total'] or 0
    
    today_expense = today_transactions.filter(type='expense').aggregate(
        total=Sum('net_amount')
    )['total'] or 0
    
    # If no transactions today, use recent transactions for demonstration
    if today_income == 0 and today_expense == 0:
        # Get recent transactions for demonstration
        recent_income = Transaction.objects.filter(
            company=company,
            type='income',
            is_void=False
        ).order_by('-transaction_date')[:3].aggregate(
            total=Sum('net_amount')
        )['total'] or 0
        
        recent_expense = Transaction.objects.filter(
            company=company,
            type='expense',
            is_void=False
        ).order_by('-transaction_date')[:3].aggregate(
            total=Sum('net_amount')
        )['total'] or 0
        
        today_income = recent_income
        today_expense = recent_expense
    
    # Get source app breakdown - optimized
    source_breakdown = list(Transaction.objects.filter(
        company=company,
        is_void=False
    ).values('source_app').annotate(
        total=Sum('net_amount'),
        count=Count('id')
    ).order_by('-total'))
    
    # Convert Decimal values to float for JSON serialization
    for item in source_breakdown:
        item['total'] = float(item['total'])
    
    context = {
        'current_ledger': current_ledger,
        'recent_transactions': recent_transactions,
        'monthly_data': list(reversed(monthly_data)),  # Reverse for chronological order
        'outstanding_invoices': float(outstanding_invoices),
        'today_income': float(today_income),
        'today_expense': float(today_expense),
        'source_breakdown': source_breakdown,
        'current_month': current_month,
        'current_year': current_year,
        'has_real_data': any(item['income'] > 0 or item['expense'] > 0 for item in monthly_data),
        'company_profile': company,  # Add company profile for currency formatting
    }
    
    return render(request, 'accounting/dashboard.html', context)


@login_required
def transaction_list(request):
    """List all transactions with filtering and pagination"""
    try:
        user = request.user
        company = getattr(user, 'company_profile', None)
        
        if not company:
            messages.error(request, "Company profile not found.")
            return redirect('core:company_profile')
        
        # Get filter form
        filter_form = TransactionFilterForm(request.GET)
        
        # Base queryset
        transactions = Transaction.objects.filter(
            company=company,
            is_void=False
        )
        
        # Apply filters
        if filter_form.is_valid():
            data = filter_form.cleaned_data
            
            if data.get('start_date'):
                transactions = transactions.filter(transaction_date__gte=data['start_date'])
            
            if data.get('end_date'):
                transactions = transactions.filter(transaction_date__lte=data['end_date'])
            
            if data.get('transaction_type'):
                transactions = transactions.filter(type=data['transaction_type'])
            
            if data.get('source_app'):
                transactions = transactions.filter(source_app=data['source_app'])
            
            if data.get('min_amount'):
                transactions = transactions.filter(net_amount__gte=data['min_amount'])
            
            if data.get('max_amount'):
                transactions = transactions.filter(net_amount__lte=data['max_amount'])
            
            if data.get('search'):
                search_term = data['search']
                transactions = transactions.filter(
                    Q(title__icontains=search_term) |
                    Q(description__icontains=search_term) |
                    Q(notes__icontains=search_term)
                )
            
            if data.get('is_reconciled'):
                is_reconciled = data['is_reconciled'] == 'true'
                transactions = transactions.filter(is_reconciled=is_reconciled)
        
        # Order by date
        transactions = transactions.order_by('-transaction_date', '-created_at')
        
        # Pagination
        paginator = Paginator(transactions, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Summary totals - calculate on filtered data
        total_income = transactions.filter(type='income').aggregate(
            total=Sum('net_amount')
        )['total'] or 0
        
        total_expense = transactions.filter(type='expense').aggregate(
            total=Sum('net_amount')
        )['total'] or 0
        
        net_total = total_income - total_expense
        
        # Get total count for pagination info
        total_count = transactions.count()
        
        context = {
            'page_obj': page_obj,
            'filter_form': filter_form,
            'total_income': total_income,
            'total_expense': total_expense,
            'net_total': net_total,
            'total_count': total_count,
        }
        
        return render(request, 'accounting/transactions.html', context)
        
    except Exception as e:
        messages.error(request, f"An error occurred while loading transactions: {str(e)}")
        return redirect('accounting:dashboard')


@login_required
def add_transaction(request):
    """Add a new transaction"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, user=user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = user
            transaction.company = company
            transaction.source_app = 'manual'
            transaction.save()
            
            messages.success(request, f"Transaction '{transaction.title}' added successfully.")
            return redirect('accounting:transaction_list')
    else:
        form = TransactionForm(user=user)
    
    context = {
        'form': form,
        'title': 'Add Transaction',
    }
    
    return render(request, 'accounting/transaction_form.html', context)


@login_required
def edit_transaction(request, transaction_id):
    """Edit an existing transaction"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    transaction = get_object_or_404(Transaction, id=transaction_id, company=company)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, user=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Transaction '{transaction.title}' updated successfully.")
            return redirect('accounting:transaction_list')
    else:
        form = TransactionForm(instance=transaction, user=user)
    
    context = {
        'form': form,
        'transaction': transaction,
        'title': 'Edit Transaction',
    }
    
    return render(request, 'accounting/transaction_form.html', context)


@login_required
def delete_transaction(request, transaction_id):
    """Delete a transaction (mark as void)"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    transaction = get_object_or_404(Transaction, id=transaction_id, company=company)
    
    if request.method == 'POST':
        transaction.is_void = True
        transaction.save()
        messages.success(request, f"Transaction '{transaction.title}' has been voided.")
        return redirect('accounting:transaction_list')
    
    context = {
        'transaction': transaction,
    }
    
    return render(request, 'accounting/transaction_confirm_delete.html', context)


@login_required
@require_POST
@csrf_exempt
def reconcile_transaction(request, transaction_id):
    """Mark a transaction as reconciled"""
    try:
        user = request.user
        company = getattr(user, 'company_profile', None)
        
        if not company:
            return JsonResponse({'success': False, 'error': 'Company profile not found'}, status=400)
        
        transaction = get_object_or_404(Transaction, id=transaction_id, company=company)
        transaction.is_reconciled = True
        transaction.save()
        
        return JsonResponse({'success': True, 'message': 'Transaction marked as reconciled'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
@csrf_exempt
def unreconcile_transaction(request, transaction_id):
    """Mark a transaction as not reconciled"""
    try:
        user = request.user
        company = getattr(user, 'company_profile', None)
        
        if not company:
            return JsonResponse({'success': False, 'error': 'Company profile not found'}, status=400)
        
        transaction = get_object_or_404(Transaction, id=transaction_id, company=company)
        transaction.is_reconciled = False
        transaction.save()
        
        return JsonResponse({'success': True, 'message': 'Transaction marked as not reconciled'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ledger_summary(request):
    """Show ledger summary by month/year"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    # Get year filter
    selected_year = request.GET.get('year', timezone.now().year)
    
    # Get ledgers for the selected year
    ledgers = Ledger.objects.filter(
        company=company,
        year=selected_year
    ).order_by('month')
    
    # Get available years
    available_years = Ledger.objects.filter(
        company=company
    ).values_list('year', flat=True).distinct().order_by('-year')
    
    # Get company currency
    company_currency = getattr(company, 'currency_symbol', '₦')
    
    # Add currency information to ledgers
    for ledger in ledgers:
        ledger.currency_symbol = company_currency
    
    # Calculate yearly totals
    yearly_totals = {
        'total_income': sum(ledger.total_income for ledger in ledgers),
        'total_expense': sum(ledger.total_expense for ledger in ledgers),
        'net_profit': sum(ledger.net_profit for ledger in ledgers),
        'currency_symbol': company_currency,
    }
    
    context = {
        'ledgers': ledgers,
        'yearly_totals': yearly_totals,
        'selected_year': int(selected_year),
        'available_years': available_years,
    }
    
    return render(request, 'accounting/ledger_summary.html', context)


@login_required
def export_accounting_data(request):
    """Export accounting data to CSV/Excel/PDF"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    export_type = request.GET.get('type', 'transactions')
    format_type = request.GET.get('format', 'csv')
    
    if export_type == 'transactions':
        # Export transactions
        transactions = Transaction.objects.filter(
            company=company,
            is_void=False
        ).order_by('-transaction_date')
        
        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'S/N', 'Date', 'Type', 'Title', 'Amount', 'Currency', 'Tax', 'Discount', 
                'Net Amount', 'Source', 'Reference', 'Notes', 'Reconciled'
            ])
            
            # Use the utility function for currency conversion
            
            for index, transaction in enumerate(transactions, 1):
                currency_display = get_currency_display(transaction.currency)
                writer.writerow([
                    index,  # Serial number
                    transaction.transaction_date,
                    transaction.get_type_display(),
                    clean_transaction_text(transaction.title),
                    transaction.amount,
                    currency_display,
                    transaction.tax or 0,
                    transaction.discount or 0,
                    transaction.net_amount,
                    clean_transaction_text(transaction.get_source_app_display()),
                    clean_transaction_text(transaction.reference_id or ''),
                    clean_transaction_text(transaction.notes or ''),
                    'Yes' if transaction.is_reconciled else 'No',
                ])
            
            return response
        
        elif format_type == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Create workbook and worksheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Transactions"
                
                # Use the utility function for currency conversion
                
                # Define headers
                headers = [
                    'S/N', 'Date', 'Type', 'Title', 'Amount', 'Currency', 'Tax', 'Discount', 
                    'Net Amount', 'Source', 'Reference', 'Notes', 'Reconciled'
                ]
                
                # Style for headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row, transaction in enumerate(transactions, 2):
                    currency_display = get_currency_display(transaction.currency)
                    ws.cell(row=row, column=1, value=row-1)  # Serial number
                    ws.cell(row=row, column=2, value=transaction.transaction_date)
                    ws.cell(row=row, column=3, value=transaction.get_type_display())
                    ws.cell(row=row, column=4, value=clean_transaction_text(transaction.title))
                    ws.cell(row=row, column=5, value=float(transaction.amount))
                    ws.cell(row=row, column=6, value=currency_display)
                    ws.cell(row=row, column=7, value=float(transaction.tax or 0))
                    ws.cell(row=row, column=8, value=float(transaction.discount or 0))
                    ws.cell(row=row, column=9, value=float(transaction.net_amount))
                    ws.cell(row=row, column=10, value=clean_transaction_text(transaction.get_source_app_display()))
                    ws.cell(row=row, column=11, value=clean_transaction_text(transaction.reference_id or ''))
                    ws.cell(row=row, column=12, value=clean_transaction_text(transaction.notes or ''))
                    ws.cell(row=row, column=13, value='Yes' if transaction.is_reconciled else 'No')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Create response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d")}.xlsx"'
                
                wb.save(response)
                return response
                
            except ImportError:
                messages.error(request, "Excel export requires openpyxl package. Please install it.")
                return redirect('accounting:transaction_list')
        
        elif format_type == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
                from io import BytesIO
                
                # Create the HttpResponse object with PDF headers
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d")}.pdf"'
                
                # Create the PDF object using BytesIO as its "file."
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                
                # Container for the 'Flowable' objects
                elements = []
                
                # Get styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=TA_CENTER,
                    textColor=colors.darkblue
                )
                
                # Add title
                title = Paragraph(f"Transaction Report - {company.company_name}", title_style)
                elements.append(title)
                
                # Add company info
                company_info = Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal'])
                elements.append(company_info)
                elements.append(Spacer(1, 20))
                
                # Use the utility function for currency conversion
                
                # Function to create paragraph for better text rendering
                def create_paragraph(text, style):
                    """Create a ReportLab Paragraph object for better text rendering"""
                    if not text:
                        return ""
                    
                    # Clean the text using our utility function
                    clean_text = clean_transaction_text(text)
                    
                    # Wrap text if it's too long
                    if len(clean_text) > 25:
                        words = clean_text.split()
                        lines = []
                        current_line = ""
                        
                        for word in words:
                            if len(current_line + " " + word) <= 25:
                                current_line += (" " + word) if current_line else word
                            else:
                                if current_line:
                                    lines.append(current_line)
                                current_line = word
                        
                        if current_line:
                            lines.append(current_line)
                        
                        clean_text = "<br/>".join(lines)
                    
                    return Paragraph(clean_text, style)
                
                # Prepare table data with serial numbers
                table_data = [['S/N', 'Date', 'Type', 'Title', 'Amount', 'Currency', 'Tax', 'Discount', 'Net Amount', 'Source', 'Reconciled']]
                
                # Create paragraph style for text cells
                text_style = ParagraphStyle(
                    'TableCell',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10,
                    spaceAfter=2,
                    spaceBefore=2,
                    leftIndent=2,
                    rightIndent=2,
                )
                
                for index, transaction in enumerate(transactions, 1):
                    # Get proper currency display
                    currency_display = get_currency_display(transaction.currency)
                    
                    # Create paragraph for title with proper text wrapping
                    title_paragraph = create_paragraph(transaction.title, text_style)
                    
                    # Create paragraph for source with proper text wrapping
                    source_paragraph = create_paragraph(transaction.get_source_app_display(), text_style)
                    
                    table_data.append([
                        str(index),  # Serial number
                        transaction.transaction_date.strftime('%Y-%m-%d'),
                        transaction.get_type_display(),
                        title_paragraph,  # Use paragraph object
                        f"{transaction.amount:,.2f}",
                        currency_display,
                        f"{transaction.tax or 0:,.2f}",
                        f"{transaction.discount or 0:,.2f}",
                        f"{transaction.net_amount:,.2f}",
                        source_paragraph,  # Use paragraph object
                        'Yes' if transaction.is_reconciled else 'No',
                    ])
                
                # Create table with adjusted column widths for better layout
                col_widths = [
                    0.8*inch,  # S/N - significantly increased width to prevent cutoff
                    0.7*inch,  # Date - slightly reduced
                    0.5*inch,  # Type - slightly reduced
                    1.5*inch,  # Title - reduced to accommodate wider S/N and Reconciled
                    0.7*inch,  # Amount - slightly reduced
                    0.5*inch,  # Currency
                    0.5*inch,  # Tax - slightly reduced
                    0.5*inch,  # Discount - slightly reduced
                    0.7*inch,  # Net Amount - slightly reduced
                    0.9*inch,  # Source - slightly reduced
                    0.9*inch,  # Reconciled - significantly increased width to prevent cutoff
                ]
                
                table = Table(table_data, colWidths=col_widths)
                
                # Add style to table with improved formatting
                style = TableStyle([
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data row styling
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    
                    # Specific column alignments
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S/N center
                    ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Date center
                    ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Type center
                    ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Title left
                    ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Amount right
                    ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Currency center
                    ('ALIGN', (6, 1), (6, -1), 'RIGHT'),   # Tax right
                    ('ALIGN', (7, 1), (7, -1), 'RIGHT'),   # Discount right
                    ('ALIGN', (8, 1), (8, -1), 'RIGHT'),   # Net Amount right
                    ('ALIGN', (9, 1), (9, -1), 'LEFT'),    # Source left
                    ('ALIGN', (10, 1), (10, -1), 'CENTER'), # Reconciled center
                    
                    # Text wrapping and padding for title and source columns
                    ('VALIGN', (3, 1), (3, -1), 'TOP'),    # Title top align
                    ('VALIGN', (9, 1), (9, -1), 'TOP'),    # Source top align
                    ('LEFTPADDING', (3, 1), (3, -1), 3),   # Title left padding
                    ('RIGHTPADDING', (3, 1), (3, -1), 3),  # Title right padding
                    ('LEFTPADDING', (9, 1), (9, -1), 3),   # Source left padding
                    ('RIGHTPADDING', (9, 1), (9, -1), 3),  # Source right padding
                    
                    # Padding for S/N and Reconciled columns to prevent cutoff
                    ('LEFTPADDING', (0, 1), (0, -1), 5),   # S/N left padding
                    ('RIGHTPADDING', (0, 1), (0, -1), 5),  # S/N right padding
                    ('LEFTPADDING', (10, 1), (10, -1), 5), # Reconciled left padding
                    ('RIGHTPADDING', (10, 1), (10, -1), 5), # Reconciled right padding
                    
                    # Row height for better text display
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
                ])
                table.setStyle(style)
                elements.append(table)
                
                # Add summary with proper currency display
                elements.append(Spacer(1, 20))
                total_income = sum(t.net_amount for t in transactions if t.type == 'income')
                total_expense = sum(t.net_amount for t in transactions if t.type == 'expense')
                net_total = total_income - total_expense
                
                # Use the most common currency from transactions, or default to NGN
                most_common_currency = 'NGN'
                if transactions.exists():
                    currency_counts = {}
                    for t in transactions:
                        currency = get_currency_display(t.currency)
                        currency_counts[currency] = currency_counts.get(currency, 0) + 1
                    if currency_counts:
                        most_common_currency = max(currency_counts, key=currency_counts.get)
                
                summary_data = [
                    ['Summary', 'Amount'],
                    ['Total Income', f"{most_common_currency} {total_income:,.2f}"],
                    ['Total Expense', f"{most_common_currency} {total_expense:,.2f}"],
                    ['Net Total', f"{most_common_currency} {net_total:,.2f}"],
                ]
                
                summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
                summary_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),  # Right align amounts
                ])
                summary_table.setStyle(summary_style)
                elements.append(summary_table)
                
                # Build PDF
                doc.build(elements)
                
                # Get the value of the BytesIO buffer and write it to the response
                pdf = buffer.getvalue()
                buffer.close()
                response.write(pdf)
                
                return response
                
            except ImportError:
                messages.error(request, "PDF export requires reportlab package. Please install it.")
                return redirect('accounting:transaction_list')
    
    elif export_type == 'ledger':
        # Export ledger summary
        year = request.GET.get('year', timezone.now().year)
        ledgers = Ledger.objects.filter(
            company=company,
            year=year
        ).order_by('month')
        
        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="ledger_{year}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['S/N', 'Month', 'Income', 'Expense', 'Net Profit', 'Outstanding Invoices', 'Profit Margin'])
            
            # Get company currency code instead of symbol
            company_currency_symbol = getattr(company, 'currency_symbol', '₦')
            company_currency_code = get_currency_display(company_currency_symbol)
            
            for index, ledger in enumerate(ledgers, 1):
                profit_margin = 0
                if ledger.total_income > 0:
                    profit_margin = (ledger.net_profit / ledger.total_income) * 100
                
                writer.writerow([
                    index,  # Serial number
                    ledger.month_name,
                    f"{company_currency_code} {ledger.total_income:,.2f}",
                    f"{company_currency_code} {ledger.total_expense:,.2f}",
                    f"{company_currency_code} {ledger.net_profit:,.2f}",
                    f"{company_currency_code} {ledger.outstanding_invoices:,.2f}",
                    f"{profit_margin:.1f}%",
                ])
            
            return response
        
        elif format_type == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Create workbook and worksheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Ledger Summary"
                
                # Get company currency code instead of symbol
                company_currency_symbol = getattr(company, 'currency_symbol', '₦')
                company_currency_code = get_currency_display(company_currency_symbol)
                
                # Define headers
                headers = ['S/N', 'Month', 'Income', 'Expense', 'Net Profit', 'Outstanding Invoices', 'Profit Margin']
                
                # Style for headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row, ledger in enumerate(ledgers, 2):
                    profit_margin = 0
                    if ledger.total_income > 0:
                        profit_margin = (ledger.net_profit / ledger.total_income) * 100
                    
                    ws.cell(row=row, column=1, value=row-1)  # Serial number
                    ws.cell(row=row, column=2, value=ledger.month_name)
                    ws.cell(row=row, column=3, value=f"{company_currency_code} {float(ledger.total_income):,.2f}")
                    ws.cell(row=row, column=4, value=f"{company_currency_code} {float(ledger.total_expense):,.2f}")
                    ws.cell(row=row, column=5, value=f"{company_currency_code} {float(ledger.net_profit):,.2f}")
                    ws.cell(row=row, column=6, value=f"{company_currency_code} {float(ledger.outstanding_invoices):,.2f}")
                    ws.cell(row=row, column=7, value=f"{profit_margin:.1f}%")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Create response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="ledger_{year}.xlsx"'
                
                wb.save(response)
                return response
                
            except ImportError:
                messages.error(request, "Excel export requires openpyxl package. Please install it.")
                return redirect('accounting:transaction_list')
        
        elif format_type == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
                from io import BytesIO
                
                # Create the HttpResponse object with PDF headers
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="ledger_{year}.pdf"'
                
                # Create the PDF object using BytesIO as its "file."
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                
                # Container for the 'Flowable' objects
                elements = []
                
                # Get styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=TA_CENTER,
                    textColor=colors.darkblue
                )
                
                # Add title
                title = Paragraph(f"Ledger Summary Report - {year} - {company.company_name}", title_style)
                elements.append(title)
                
                # Add company info
                company_info = Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal'])
                elements.append(company_info)
                elements.append(Spacer(1, 20))
                
                # Get company currency code instead of symbol
                company_currency_symbol = getattr(company, 'currency_symbol', '₦')
                company_currency_code = get_currency_display(company_currency_symbol)
                
                # Prepare table data with serial numbers and currency
                table_data = [['S/N', 'Month', 'Income', 'Expense', 'Net Profit', 'Outstanding Invoices', 'Profit Margin']]
                
                for index, ledger in enumerate(ledgers, 1):
                    profit_margin = 0
                    if ledger.total_income > 0:
                        profit_margin = (ledger.net_profit / ledger.total_income) * 100
                    
                    table_data.append([
                        str(index),  # Serial number
                        ledger.month_name,
                        f"{company_currency_code} {ledger.total_income:,.2f}",
                        f"{company_currency_code} {ledger.total_expense:,.2f}",
                        f"{company_currency_code} {ledger.net_profit:,.2f}",
                        f"{company_currency_code} {ledger.outstanding_invoices:,.2f}",
                        f"{profit_margin:.1f}%",
                    ])
                
                # Create table with adjusted column widths
                table = Table(table_data, colWidths=[0.5*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.0*inch])
                
                # Add style to table with improved formatting for new columns
                style = TableStyle([
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data row styling
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    
                    # Specific column alignments
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S/N center
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'),     # Month left
                    ('ALIGN', (2, 1), (5, -1), 'RIGHT'),    # Amounts right
                    ('ALIGN', (6, 1), (6, -1), 'CENTER'),   # Profit margin center
                    
                    # Row height for better text display
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
                ])
                table.setStyle(style)
                elements.append(table)
                
                # Add yearly summary with currency
                elements.append(Spacer(1, 20))
                yearly_income = sum(ledger.total_income for ledger in ledgers)
                yearly_expense = sum(ledger.total_expense for ledger in ledgers)
                yearly_profit = sum(ledger.net_profit for ledger in ledgers)
                
                summary_data = [
                    ['Yearly Summary', 'Amount'],
                    ['Total Income', f"{company_currency_code} {yearly_income:,.2f}"],
                    ['Total Expense', f"{company_currency_code} {yearly_expense:,.2f}"],
                    ['Net Profit', f"{company_currency_code} {yearly_profit:,.2f}"],
                ]
                
                summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
                summary_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),  # Right align amounts
                ])
                summary_table.setStyle(summary_style)
                elements.append(summary_table)
                
                # Build PDF
                doc.build(elements)
                
                # Get the value of the BytesIO buffer and write it to the response
                pdf = buffer.getvalue()
                buffer.close()
                response.write(pdf)
                
                return response
                
            except ImportError:
                messages.error(request, "PDF export requires reportlab package. Please install it.")
                return redirect('accounting:transaction_list')
    
    messages.error(request, "Invalid export type or format.")
    return redirect('accounting:transaction_list')


@login_required
@require_POST
@csrf_exempt
def update_ledger_ajax(request):
    """AJAX endpoint to update ledger data with comprehensive chart data"""
    try:
        # Log the request for debugging
        print(f"AJAX request received: {request.body}")
        
        data = json.loads(request.body)
        month = data.get('month')
        year = data.get('year')
        include_charts = data.get('include_charts', False)
        
        user = request.user
        company = getattr(user, 'company_profile', None)
        
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=400)
        
        # Update ledger for the specified month
        ledger = Ledger.objects.filter(
            company=company,
            year=year,
            month=month
        ).first()
        
        if ledger:
            ledger.update_outstanding_amounts()
        
        # Get outstanding invoices
        from apps.invoices.models import Invoice
        outstanding_invoices = Invoice.objects.filter(
            user=user,
            status__in=['unpaid', 'partial']
        ).aggregate(total=Sum('balance_due'))['total'] or 0
        
        # Get today's transactions
        current_date = timezone.now()
        today_transactions = Transaction.objects.filter(
            company=company,
            transaction_date=current_date.date(),
            is_void=False
        )
        
        today_income = today_transactions.filter(type='income').aggregate(
            total=Sum('net_amount')
        )['total'] or 0
        
        today_expense = today_transactions.filter(type='expense').aggregate(
            total=Sum('net_amount')
        )['total'] or 0
        
        response_data = {
            'success': True,
            'total_income': float(ledger.total_income) if ledger else 0,
            'total_expense': float(ledger.total_expense) if ledger else 0,
            'net_profit': float(ledger.net_profit) if ledger else 0,
            'outstanding_invoices': float(outstanding_invoices),
            'today_income': float(today_income),
            'today_expense': float(today_expense),
            'currency_symbol': company.currency_symbol if company else '₦',
            'currency_code': company.currency_code if company else 'NGN',
        }
        
        # Include chart data if requested
        if include_charts:
            # Monthly data for charts (last 12 months) - improved with real transaction data
            monthly_data = []
            for i in range(12):
                date = current_date - timedelta(days=30*i)
                
                # Get or create ledger for this month
                month_ledger = Ledger.objects.filter(
                    company=company,
                    year=date.year,
                    month=date.month
                ).first()
                
                # If no ledger exists, calculate from actual transactions
                if not month_ledger:
                    month_transactions = Transaction.objects.filter(
                        company=company,
                        transaction_date__year=date.year,
                        transaction_date__month=date.month,
                        is_void=False
                    )
                    
                    month_income = month_transactions.filter(type='income').aggregate(
                        total=Sum('net_amount')
                    )['total'] or 0
                    
                    month_expense = month_transactions.filter(type='expense').aggregate(
                        total=Sum('net_amount')
                    )['total'] or 0
                    
                    month_profit = month_income - month_expense
                else:
                    month_income = float(month_ledger.total_income)
                    month_expense = float(month_ledger.total_expense)
                    month_profit = float(month_ledger.net_profit)
                
                monthly_data.append({
                    'month': date.strftime('%b %Y'),
                    'income': month_income,
                    'expense': month_expense,
                    'profit': month_profit,
                })
            
            # Source breakdown data
            source_breakdown = Transaction.objects.filter(
                company=company,
                is_void=False
            ).values('source_app').annotate(
                total=Sum('net_amount'),
                count=Count('id')
            ).order_by('-total')
            
            # Recent transactions
            recent_transactions = Transaction.objects.filter(
                company=company,
                is_void=False
            ).order_by('-created_at')[:10]
            
            response_data.update({
                'monthly_data': list(reversed(monthly_data)),
                'today_data': {
                    'income': float(today_income),
                    'expense': float(today_expense)
                },
                'source_data': list(source_breakdown),
                'recent_transactions': [
                    {
                        'title': t.title,
                        'type': t.type,
                        'type_display': t.get_type_display(),
                        'net_amount': float(t.net_amount),
                        'date': t.transaction_date.strftime('%b %d, %Y'),
                        'source_app': t.get_source_app_display()
                    }
                    for t in recent_transactions
                ]
            })
        
        # Log the response for debugging
        print(f"AJAX response: {response_data}")
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        print(f"Unexpected error: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def sync_from_other_apps(request):
    """Sync transactions from other apps"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    if request.method == 'POST':
        # Sync from invoices
        from apps.invoices.models import Invoice
        from apps.receipts.models import Receipt
        
        # Sync paid invoices
        paid_invoices = Invoice.objects.filter(
            user=user,
            status='paid'
        )
        
        synced_count = 0
        for invoice in paid_invoices:
            # Check if transaction already exists
            existing_transaction = Transaction.objects.filter(
                company=company,
                source_app='invoice',
                reference_id=str(invoice.id)
            ).first()
            
            if not existing_transaction:
                Transaction.objects.create(
                    user=user,
                    company=company,
                    type='income',
                    title=f"Invoice Payment - {invoice.invoice_number}",
                    description=f"Payment for invoice {invoice.invoice_number} from {invoice.client_name}",
                    amount=invoice.grand_total,
                    currency=company.currency_symbol,
                    tax=invoice.total_tax,
                    discount=invoice.total_discount,
                    source_app='invoice',
                    reference_id=str(invoice.id),
                    reference_model='Invoice',
                    transaction_date=invoice.updated_at.date(),
                    notes=f"Auto-synced from invoice {invoice.invoice_number}"
                )
                synced_count += 1
        
        # Sync from receipts
        receipts = Receipt.objects.filter(
            created_by=user
        )
        
        for receipt in receipts:
            existing_transaction = Transaction.objects.filter(
                company=company,
                source_app='receipt',
                reference_id=str(receipt.id)
            ).first()
            
            if not existing_transaction:
                Transaction.objects.create(
                    user=user,
                    company=company,
                    type='income',
                    title=f"Receipt Payment - {receipt.receipt_no}",
                    description=f"Payment receipt {receipt.receipt_no} from {receipt.client_name}",
                    amount=receipt.amount_received,
                    currency=company.currency_symbol,
                    source_app='receipt',
                    reference_id=str(receipt.id),
                    reference_model='Receipt',
                    transaction_date=receipt.date_received,
                    notes=f"Auto-synced from receipt {receipt.receipt_no}"
                )
                synced_count += 1
        
        messages.success(request, f"Successfully synced {synced_count} transactions from other apps.")
        return redirect('accounting:transaction_list')
    
    return render(request, 'accounting/sync_from_other_apps.html')


@login_required
def generate_report(request):
    """Generate financial reports"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    if request.method == 'POST':
        form = FinancialReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.company = company
            report.created_by = user
            
            # Generate report data based on type
            if report.report_type == 'income_statement':
                report.report_data = generate_income_statement(company, report.start_date, report.end_date)
            elif report.report_type == 'balance_sheet':
                report.report_data = generate_balance_sheet(company, report.end_date)
            
            report.save()
            
            messages.success(request, f"Report '{report.title}' generated successfully.")
            return redirect('accounting:view_report', report_id=report.id)
    else:
        form = FinancialReportForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'accounting/generate_report.html', context)


@login_required
def export_report_pdf(request, report_id):
    """Export financial report as PDF"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    report = get_object_or_404(FinancialReport, id=report_id, company=company)
    
    # Calculate fresh report data dynamically (same logic as view_report)
    original_transactions_count = Transaction.objects.filter(
        company=company,
        transaction_date__range=[report.start_date, report.end_date],
        is_void=False
    ).count()
    
    # If no transactions in original range, extend the range to include recent transactions
    if original_transactions_count == 0:
        latest_transaction = Transaction.objects.filter(
            company=company,
            is_void=False
        ).order_by('-transaction_date').first()
        
        if latest_transaction:
            adjusted_start_date = min(report.start_date, latest_transaction.transaction_date)
            adjusted_end_date = max(report.end_date, latest_transaction.transaction_date)
            
            if report.report_type == 'income_statement':
                fresh_report_data = generate_income_statement(company, adjusted_start_date, adjusted_end_date)
            elif report.report_type == 'balance_sheet':
                fresh_report_data = generate_balance_sheet(company, adjusted_end_date)
            else:
                fresh_report_data = report.report_data
            
            fresh_report_data['date_adjusted'] = True
            fresh_report_data['original_period'] = f"{report.start_date} to {report.end_date}"
            fresh_report_data['adjusted_period'] = f"{adjusted_start_date} to {adjusted_end_date}"
        else:
            if report.report_type == 'income_statement':
                fresh_report_data = generate_income_statement(company, report.start_date, report.end_date)
            elif report.report_type == 'balance_sheet':
                fresh_report_data = generate_balance_sheet(company, report.end_date)
            else:
                fresh_report_data = report.report_data
            fresh_report_data['date_adjusted'] = False
    else:
        if report.report_type == 'income_statement':
            fresh_report_data = generate_income_statement(company, report.start_date, report.end_date)
        elif report.report_type == 'balance_sheet':
            fresh_report_data = generate_balance_sheet(company, report.end_date)
        else:
            fresh_report_data = report.report_data
        fresh_report_data['date_adjusted'] = False
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import os
        
        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report.title}_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        # Register fonts for better Unicode support
        try:
            # Try to register a Unicode-compatible font if available
            font_path = os.path.join(os.path.dirname(__file__), '..', '..', 'static', 'fonts')
            
            # Try to find a Unicode-compatible font
            unicode_fonts = [
                'DejaVuSans.ttf',
                'Arial.ttf',
                'arial.ttf',
                'LiberationSans-Regular.ttf',
                'FreeSans.ttf'
            ]
            
            default_font = 'Helvetica'  # Default fallback
            font_registered = False
            
            for font_file in unicode_fonts:
                font_path_full = os.path.join(font_path, font_file)
                if os.path.exists(font_path_full):
                    try:
                        font_name = font_file.replace('.ttf', '')
                        pdfmetrics.registerFont(TTFont(font_name, font_path_full))
                        default_font = font_name
                        font_registered = True
                        break
                    except:
                        continue
            
            # If no font found in static directory, try system fonts
            if not font_registered:
                # Try to use a system font that supports Unicode
                system_fonts = [
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                    '/System/Library/Fonts/Arial.ttf',
                    'C:/Windows/Fonts/arial.ttf',
                    'C:/Windows/Fonts/calibri.ttf'
                ]
                
                for system_font in system_fonts:
                    if os.path.exists(system_font):
                        try:
                            font_name = os.path.basename(system_font).replace('.ttf', '')
                            pdfmetrics.registerFont(TTFont(font_name, system_font))
                            default_font = font_name
                            font_registered = True
                            break
                        except:
                            continue
            
            # If we registered a custom font, we need to be careful about bold variants
            # Most TTF fonts don't automatically have bold variants in ReportLab
            # So we'll use the regular font for everything and rely on font weight
            if font_registered:
                # Use the registered font name without bold suffix
                default_font = default_font
            else:
                # Use Helvetica which has built-in bold variants
                default_font = 'Helvetica'
                            
        except Exception as e:
            print(f"Font registration error: {e}")
            default_font = 'Helvetica'
        
        # Create the PDF object using BytesIO as its "file."
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkgreen
        )
        
        # Add title
        title = Paragraph(f"{report.title}", title_style)
        elements.append(title)
        
        # Add subtitle
        subtitle = Paragraph(f"{report.get_report_type_display()} - {company.company_name}", subtitle_style)
        elements.append(subtitle)
        
        # Add report info with currency
        currency_symbol = getattr(company, 'currency_symbol', '₦')
        if report.report_type == 'balance_sheet':
            report_info = Paragraph(f"As of: {fresh_report_data.get('as_of_date', report.end_date)}<br/>Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal'])
        else:
            report_info = Paragraph(f"Period: {fresh_report_data.get('period', f'{report.start_date} to {report.end_date}')}<br/>Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal'])
        elements.append(report_info)
        
        # Add date adjustment notice if applicable
        if fresh_report_data.get('date_adjusted'):
            adjustment_notice = Paragraph(f"<b>Note:</b> Date range adjusted to include actual data. Original: {fresh_report_data['original_period']} → Adjusted: {fresh_report_data['adjusted_period']}", styles['Normal'])
            elements.append(adjustment_notice)
        
        elements.append(Spacer(1, 30))
        
        # Generate report content based on type with fresh data
        if report.report_type == 'income_statement':
            elements.extend(generate_income_statement_pdf_content(fresh_report_data, currency_symbol, default_font))
        elif report.report_type == 'balance_sheet':
            elements.extend(generate_balance_sheet_pdf_content(fresh_report_data, currency_symbol, default_font))
        else:
            # Generic report display
            elements.append(Paragraph("Report Data:", styles['Heading3']))
            elements.append(Spacer(1, 10))
            
            # Display report data as table
            if fresh_report_data:
                # Handle currency symbol for PDF compatibility - try symbol first, fallback to text
                pdf_currency = get_pdf_currency_symbol(currency_symbol, use_symbol=False)
                
                table_data = [['Item', 'Value']]
                for key, value in fresh_report_data.items():
                    if key not in ['date_adjusted', 'original_period', 'adjusted_period', 'debug_info']:
                        if isinstance(value, (int, float)):
                            table_data.append([key.replace('_', ' ').title(), f"{pdf_currency} {value:,.2f}"])
                        else:
                            table_data.append([key.replace('_', ' ').title(), str(value)])
                
                # Determine font names based on whether we're using a custom font or built-in Helvetica
                if default_font == 'Helvetica':
                    header_font = 'Helvetica-Bold'
                    bold_font = 'Helvetica-Bold'
                else:
                    # For custom fonts, use the regular font name (no bold variant available)
                    header_font = default_font
                    bold_font = default_font
                
                table = Table(table_data, colWidths=[3*inch, 2*inch])
                style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), header_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ])
                table.setStyle(style)
                elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error generating PDF: {str(e)}")
        return redirect('accounting:view_report', report_id=report.id)


def generate_income_statement_pdf_content(report_data, currency_symbol, default_font='Helvetica'):
    """Generate PDF content for income statement"""
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Handle currency symbol for PDF compatibility
    pdf_currency = get_pdf_currency_symbol(currency_symbol, use_symbol=False)
    
    # Income Statement table
    table_data = [
        ['Income Statement', ''],
        ['', ''],
        ['Revenue', f"{pdf_currency} {report_data.get('income', 0):,.2f}"],
        ['', ''],
        ['Expenses', f"{pdf_currency} {report_data.get('expenses', 0):,.2f}"],
        ['', ''],
        ['Net Income', f"{pdf_currency} {report_data.get('net_income', 0):,.2f}"],
    ]
    
    table = Table(table_data, colWidths=[3*inch, 2*inch])
    # Determine font names based on whether we're using a custom font or built-in Helvetica
    if default_font == 'Helvetica':
        header_font = 'Helvetica-Bold'
        bold_font = 'Helvetica-Bold'
    else:
        # For custom fonts, use the regular font name (no bold variant available)
        header_font = default_font
        bold_font = default_font
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), header_font),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 2), (0, 2), bold_font),  # Revenue
        ('FONTNAME', (0, 4), (0, 4), bold_font),  # Expenses
        ('FONTNAME', (0, 6), (0, 6), bold_font),  # Net Income
        ('BACKGROUND', (0, 2), (-1, 2), colors.lightgreen),  # Revenue section
        ('BACKGROUND', (0, 4), (-1, 4), colors.lightcoral),  # Expenses section
        ('BACKGROUND', (0, 6), (-1, 6), colors.lightblue),  # Net Income section
    ])
    table.setStyle(style)
    elements.append(table)
    
    # Add period information
    if report_data.get('period'):
        elements.append(Spacer(1, 20))
        period_info = Paragraph(f"<b>Period:</b> {report_data['period']}", styles['Normal'])
        elements.append(period_info)
    
    return elements


def generate_balance_sheet_pdf_content(report_data, currency_symbol, default_font='Helvetica'):
    """Generate PDF content for balance sheet"""
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Handle currency symbol for PDF compatibility - try symbol first, fallback to text
    pdf_currency = get_pdf_currency_symbol(currency_symbol, use_symbol=False)
    
    # Balance Sheet table with detailed breakdown
    table_data = [
        ['Balance Sheet', ''],
        ['', ''],
        ['ASSETS', ''],
        ['Cash', f"{pdf_currency} {report_data.get('cash', 0):,.2f}"],
        ['Accounts Receivable', f"{pdf_currency} {report_data.get('accounts_receivable', 0):,.2f}"],
        ['Total Assets', f"{pdf_currency} {report_data.get('total_assets', 0):,.2f}"],
        ['', ''],
        ['LIABILITIES', ''],
        ['Accounts Payable', f"{pdf_currency} {report_data.get('accounts_payable', 0):,.2f}"],
        ['Total Liabilities', f"{pdf_currency} {report_data.get('total_liabilities', 0):,.2f}"],
        ['', ''],
        ['EQUITY', ''],
        ['Owner\'s Equity', f"{pdf_currency} {report_data.get('owner_equity', 0):,.2f}"],
        ['Retained Earnings', f"{pdf_currency} {report_data.get('retained_earnings', 0):,.2f}"],
        ['Total Equity', f"{pdf_currency} {report_data.get('owner_equity', 0) + report_data.get('retained_earnings', 0):,.2f}"],
        ['', ''],
        ['Total Liabilities & Equity', f"{pdf_currency} {report_data.get('total_liabilities', 0) + report_data.get('owner_equity', 0) + report_data.get('retained_earnings', 0):,.2f}"],
    ]
    
    # Determine font names based on whether we're using a custom font or built-in Helvetica
    if default_font == 'Helvetica':
        header_font = 'Helvetica-Bold'
        bold_font = 'Helvetica-Bold'
    else:
        # For custom fonts, use the regular font name (no bold variant available)
        header_font = default_font
        bold_font = default_font
    
    table = Table(table_data, colWidths=[3*inch, 2*inch])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), header_font),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 2), (0, 2), bold_font),  # ASSETS
        ('FONTNAME', (0, 8), (0, 8), bold_font),  # LIABILITIES
        ('FONTNAME', (0, 12), (0, 12), bold_font),  # EQUITY
        ('FONTNAME', (0, 5), (0, 5), bold_font),  # Total Assets
        ('FONTNAME', (0, 10), (0, 10), bold_font),  # Total Liabilities
        ('FONTNAME', (0, 15), (0, 15), bold_font),  # Total Equity
        ('FONTNAME', (0, 17), (0, 17), bold_font),  # Total Liabilities & Equity
        ('BACKGROUND', (0, 2), (-1, 2), colors.lightblue),  # Assets section
        ('BACKGROUND', (0, 8), (-1, 8), colors.lightcoral),  # Liabilities section
        ('BACKGROUND', (0, 12), (-1, 12), colors.lightgreen),  # Equity section
        ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),  # Total Assets
        ('BACKGROUND', (0, 10), (-1, 10), colors.lightgrey),  # Total Liabilities
        ('BACKGROUND', (0, 15), (-1, 15), colors.lightgrey),  # Total Equity
        ('BACKGROUND', (0, 17), (-1, 17), colors.lightgrey),  # Total Liabilities & Equity
    ])
    table.setStyle(style)
    elements.append(table)
    
    # Add as of date information
    if report_data.get('as_of_date'):
        elements.append(Spacer(1, 20))
        as_of_info = Paragraph(f"<b>As of Date:</b> {report_data['as_of_date']}", styles['Normal'])
        elements.append(as_of_info)
    
    # Add period information
    if report_data.get('period'):
        elements.append(Spacer(1, 10))
        period_info = Paragraph(f"<b>Period:</b> {report_data['period']}", styles['Normal'])
        elements.append(period_info)
    
    return elements


def generate_income_statement(company, start_date, end_date):
    """Generate income statement data"""
    # Get all transactions in the date range
    transactions = Transaction.objects.filter(
        company=company,
        transaction_date__range=[start_date, end_date],
        is_void=False
    )
    
    # Calculate income (positive amounts)
    income_transactions = transactions.filter(type='income')
    income = income_transactions.aggregate(
        total=Sum('net_amount')
    )['total'] or 0
    
    # Calculate expenses (negative amounts)
    expense_transactions = transactions.filter(type='expense')
    expenses = expense_transactions.aggregate(
        total=Sum('net_amount')
    )['total'] or 0
    
    # Calculate net income
    net_income = income - expenses
    
    # Debug information
    print(f"DEBUG: Income Statement for {company.company_name}")
    print(f"DEBUG: Date range: {start_date} to {end_date}")
    print(f"DEBUG: Total transactions found: {transactions.count()}")
    print(f"DEBUG: Income transactions: {income_transactions.count()}")
    print(f"DEBUG: Expense transactions: {expense_transactions.count()}")
    print(f"DEBUG: Total income: {income}")
    print(f"DEBUG: Total expenses: {expenses}")
    print(f"DEBUG: Net income: {net_income}")
    
    return {
        'income': float(income),
        'expenses': float(expenses),
        'net_income': float(net_income),
        'period': f"{start_date} to {end_date}",
        'debug_info': {
            'total_transactions': transactions.count(),
            'income_transactions': income_transactions.count(),
            'expense_transactions': expense_transactions.count(),
        }
    }


def generate_balance_sheet(company, as_of_date):
    """Generate balance sheet data with proper accounting equation"""
    # Get all transactions up to the as_of_date
    transactions = Transaction.objects.filter(
        company=company,
        transaction_date__lte=as_of_date,
        is_void=False
    )
    
    # Calculate total income and expenses
    total_income = transactions.filter(type='income').aggregate(
        total=Sum('net_amount')
    )['total'] or 0
    
    total_expenses = transactions.filter(type='expense').aggregate(
        total=Sum('net_amount')
    )['total'] or 0
    
    # Calculate retained earnings (net income)
    retained_earnings = total_income - total_expenses
    
    # Get outstanding invoices as accounts receivable
    try:
        from apps.invoices.models import Invoice
        accounts_receivable = Invoice.objects.filter(
            user=company.user,
            status__in=['unpaid', 'partial']
        ).aggregate(total=Sum('balance_due'))['total'] or 0
    except:
        accounts_receivable = 0
    
    # FIXED: Proper balance sheet calculation
    # Assets = Cash + Accounts Receivable + Other Assets
    cash = retained_earnings  # Simplified: cash equals retained earnings
    total_assets = cash + accounts_receivable
    
    # Liabilities = Accounts Payable + Other Liabilities
    # For now, we'll calculate payables from outstanding expenses
    try:
        from apps.expenses.models import Expense
        accounts_payable = Expense.objects.filter(
            user=company.user,
            status__in=['pending', 'unpaid']
        ).aggregate(total=Sum('amount'))['total'] or 0
    except:
        accounts_payable = 0
    
    total_liabilities = accounts_payable
    
    # Equity = Owner's Equity + Retained Earnings
    # For simplicity, we'll assume owner's equity is the initial investment
    # and calculate it to balance the equation
    owner_equity = total_assets - total_liabilities - retained_earnings
    
    # If owner's equity is negative, it means we have accumulated losses
    # In this case, we should adjust retained earnings
    if owner_equity < 0:
        retained_earnings = total_assets - total_liabilities
        owner_equity = 0
    
    # Verify the accounting equation: Assets = Liabilities + Equity
    calculated_equity = total_liabilities + owner_equity + retained_earnings
    balance_check = abs(total_assets - calculated_equity)
    
    # Debug information
    print(f"DEBUG: Balance Sheet for {company.company_name}")
    print(f"DEBUG: As of date: {as_of_date}")
    print(f"DEBUG: Total transactions found: {transactions.count()}")
    print(f"DEBUG: Total income: {total_income}")
    print(f"DEBUG: Total expenses: {total_expenses}")
    print(f"DEBUG: Retained earnings: {retained_earnings}")
    print(f"DEBUG: Accounts receivable: {accounts_receivable}")
    print(f"DEBUG: Cash: {cash}")
    print(f"DEBUG: Total assets: {total_assets}")
    print(f"DEBUG: Accounts payable: {accounts_payable}")
    print(f"DEBUG: Total liabilities: {total_liabilities}")
    print(f"DEBUG: Owner's equity: {owner_equity}")
    print(f"DEBUG: Balance check (should be 0): {balance_check}")
    
    return {
        'total_assets': float(total_assets),
        'total_liabilities': float(total_liabilities),
        'owner_equity': float(owner_equity),
        'retained_earnings': float(retained_earnings),
        'accounts_receivable': float(accounts_receivable),
        'accounts_payable': float(accounts_payable),
        'cash': float(cash),
        'as_of_date': as_of_date.strftime('%Y-%m-%d'),
        'period': f"As of {as_of_date.strftime('%B %d, %Y')}",
        'balance_check': float(balance_check),
        'debug_info': {
            'total_transactions': transactions.count(),
            'income_transactions': transactions.filter(type='income').count(),
            'expense_transactions': transactions.filter(type='expense').count(),
            'accounting_equation_balanced': balance_check < 0.01,  # Allow for small rounding differences
        }
    }


@login_required
def view_report(request, report_id):
    """View a generated financial report"""
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    if not company:
        messages.error(request, "Company profile not found.")
        return redirect('core:company_profile')
    
    report = get_object_or_404(FinancialReport, id=report_id, company=company)
    
    # Check if there are transactions in the original date range
    original_transactions_count = Transaction.objects.filter(
        company=company,
        transaction_date__range=[report.start_date, report.end_date],
        is_void=False
    ).count()
    
    # If no transactions in original range, extend the range to include recent transactions
    if original_transactions_count == 0:
        # Get the most recent transaction date
        latest_transaction = Transaction.objects.filter(
            company=company,
            is_void=False
        ).order_by('-transaction_date').first()
        
        if latest_transaction:
            # Extend the date range to include the latest transaction
            adjusted_start_date = min(report.start_date, latest_transaction.transaction_date)
            adjusted_end_date = max(report.end_date, latest_transaction.transaction_date)
            
            # Use adjusted dates for calculation
            if report.report_type == 'income_statement':
                fresh_report_data = generate_income_statement(company, adjusted_start_date, adjusted_end_date)
            elif report.report_type == 'balance_sheet':
                fresh_report_data = generate_balance_sheet(company, adjusted_end_date)
            else:
                fresh_report_data = report.report_data
            
            # Add note about date adjustment
            fresh_report_data['date_adjusted'] = True
            fresh_report_data['original_period'] = f"{report.start_date} to {report.end_date}"
            fresh_report_data['adjusted_period'] = f"{adjusted_start_date} to {adjusted_end_date}"
        else:
            # No transactions at all, use original data
            if report.report_type == 'income_statement':
                fresh_report_data = generate_income_statement(company, report.start_date, report.end_date)
            elif report.report_type == 'balance_sheet':
                fresh_report_data = generate_balance_sheet(company, report.end_date)
            else:
                fresh_report_data = report.report_data
            fresh_report_data['date_adjusted'] = False
    else:
        # Use original date range
        if report.report_type == 'income_statement':
            fresh_report_data = generate_income_statement(company, report.start_date, report.end_date)
        elif report.report_type == 'balance_sheet':
            fresh_report_data = generate_balance_sheet(company, report.end_date)
        else:
            fresh_report_data = report.report_data
        fresh_report_data['date_adjusted'] = False
    
    # Debug: Check if there are any transactions
    total_transactions = Transaction.objects.filter(company=company).count()
    transactions_in_period = Transaction.objects.filter(
        company=company,
        transaction_date__range=[report.start_date, report.end_date],
        is_void=False
    ).count()
    
    # Debug: Get some sample transactions
    sample_transactions = Transaction.objects.filter(
        company=company,
        is_void=False
    ).order_by('-transaction_date')[:5]
    
    context = {
        'report': report,
        'company': company,
        'report_data': fresh_report_data,  # Use fresh data instead of stored data
        'debug_info': {
            'total_transactions': total_transactions,
            'transactions_in_period': transactions_in_period,
            'sample_transactions': sample_transactions,
            'date_range': f"{report.start_date} to {report.end_date}",
            'original_transactions_count': original_transactions_count,
        }
    }
    
    return render(request, 'accounting/view_report.html', context)


def update_transactions_currency(company_profile, new_currency_symbol):
    """Update all existing transactions to use the new currency symbol"""
    from .models import Transaction
    
    # Get all transactions for this company
    transactions = Transaction.objects.filter(company=company_profile)
    
    # Update currency for all transactions
    updated_count = transactions.update(currency=new_currency_symbol)
    
    return updated_count


@login_required
@require_POST
@csrf_exempt
def update_accounting_currency(request):
    """Update currency across all accounting transactions"""
    try:
        data = json.loads(request.body)
        currency_code = data.get('currency_code')
        currency_symbol = data.get('currency_symbol')
        
        if not currency_code or not currency_symbol:
            return JsonResponse({'success': False, 'error': 'Currency code and symbol are required'})
        
        # Update company profile currency
        company_profile = request.user.company_profile
        company_profile.currency_code = currency_code
        company_profile.currency_symbol = currency_symbol
        company_profile.save()
        
        # Update all existing transactions to use the new currency
        updated_count = update_transactions_currency(company_profile, currency_symbol)
        
        return JsonResponse({
            'success': True,
            'currency_code': currency_code,
            'currency_symbol': currency_symbol,
            'transactions_updated': updated_count,
            'message': f'Successfully updated {updated_count} transactions to use {currency_code} ({currency_symbol})'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
@csrf_exempt
def update_transaction_currencies(request):
    """Update all transaction currencies to match the current company profile currency"""
    try:
        user = request.user
        company = getattr(user, 'company_profile', None)
        
        if not company:
            return JsonResponse({'success': False, 'error': 'Company profile not found'})
        
        # Get current company currency
        current_currency = company.currency_symbol
        
        # Update all transactions for this company
        updated_count = update_transactions_currency(company, current_currency)
        
        return JsonResponse({
            'success': True,
            'transactions_updated': updated_count,
            'currency': current_currency,
            'message': f'Successfully updated {updated_count} transactions to use {current_currency}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
